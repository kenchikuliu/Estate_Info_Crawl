# -*- coding: utf-8 -*-
"""
Export Alibaba Guangdong auction results using the same column schema as the
JD Guangdong final detail table.

This is a schema/field-alignment step only. It does not refetch data.
"""
from __future__ import annotations

import argparse
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence

import xlsxwriter
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_ALI_INPUT = PROJECT_ROOT / r"output\ali_sf_gd_full_h5_20260611\阿里法拍房_广东_详情回填_API.xlsx"
DEFAULT_ALI_INDEX = PROJECT_ROOT / r"output\ali_sf_gd_full_h5_20260611\阿里法拍房_广东_全量索引.xlsx"
DEFAULT_ALI_DETAIL_JSONL = PROJECT_ROOT / r"output\ali_sf_gd_full_h5_20260611\阿里法拍房_广东_详情回填_API_含附件正文.jsonl"
DEFAULT_JD_SCHEMA = PROJECT_ROOT / r"output\京东法拍房_广东_详情回填_API_含附件索引.xlsx"
DEFAULT_OUTPUT = PROJECT_ROOT / r"output\ali_sf_gd_full_h5_20260611\阿里法拍房_广东_详情回填_API_京东字段对齐.xlsx"
DEFAULT_REPORT = PROJECT_ROOT / r"output\ali_sf_gd_full_h5_20260611\阿里法拍房_广东_详情回填_API_京东字段对齐.schema_verify.json"


ALIASES: Dict[str, Sequence[str]] = {
    "地址": ("完整地址", "格式化地址", "详情地址", "标的物名称", "标题"),
    "处置机构": ("处置机构", "处置法院", "处置法院/机构", "法院/处置机构", "法院", "courtName"),
    "城市ID": ("城市ID", "locationCode", "查询区域编码"),
    "时间桶开始": ("时间桶开始",),
    "时间桶结束": ("时间桶结束",),
    "评估价": ("评估价", "评估价_元", "评估价_详情_元"),
    "评估价_元": ("评估价_元", "评估价_详情_元"),
    "市场价": ("市场价", "市场价_元", "市场价_详情_元"),
    "市场价_元": ("市场价_元", "市场价_详情_元"),
    "auctionStatus": ("auctionStatus", "status", "h5BidStatus", "竞价状态", "成交状态"),
    "displayStatus": ("displayStatus", "成交状态", "竞价状态", "拍卖状态_详情", "status"),
    "是否支持保险": ("是否支持保险",),
    "是否配资服务": ("是否配资服务",),
    "labelSet": ("labelSet", "tags", "标签"),
    "loan": ("loan", "是否支持贷款", "是否有金融服务"),
    "purchaseRestriction": ("purchaseRestriction",),
    "paimaiTimes": ("paimaiTimes", "轮次筛选"),
    "skuId": ("skuId",),
    "productId": ("productId", "标的物ID"),
    "shopId": ("shopId",),
    "vendorId": ("vendorId",),
    "publishSource": ("publishSource", "平台"),
    "productImage": ("productImage", "图片链接"),
    "详情接口错误": ("详情接口错误", "详情抓取错误"),
    "auctionStatus_详情": ("auctionStatus_详情", "auctionType_详情", "拍卖状态_详情"),
    "displayStatus_详情": ("displayStatus_详情", "拍卖状态_详情", "成交状态", "竞价状态"),
    "成交确认书链接": ("成交确认书链接",),
    "是否有金融服务_详情": ("是否有金融服务_详情", "是否有金融服务"),
    "金融服务原文": ("金融服务原文", "金融机构", "最高可贷比例", "参考利率", "金融其他费用"),
    "竞买公告文本": ("竞买公告文本", "竞买公告", "拍卖公告", "标的物详情描述"),
    "附件抓取时间": ("附件抓取时间", "附件正文抓取时间", "详情抓取时间"),
    "附件抓取状态": ("附件抓取状态", "附件正文抓取状态"),
    "附件抓取错误": ("附件抓取错误", "附件正文抓取错误"),
    "附件链接": ("附件链接",),
}

PRICE_LABELS = [
    "评估价",
    "评估价值",
    "评估总价",
    "估价结果",
    "市场价值",
    "市场价",
    "询价结果",
    "参考价",
    "处置参考价",
]
PRICE_PATTERNS = {
    "评估价": re.compile(
        r"(?:评估价|评估价值|评估总价|估价结果)"
        r"(?:合计|总计|总额|为|是|：|:|\s)*"
        r"(?:人民币)?\s*([0-9][0-9,，]*(?:\.[0-9]+)?\s*(?:万元|万|元)?)"
    ),
    "市场价": re.compile(
        r"(?:市场价值|市场价|市场总价|询价结果|议价结果|参考价|处置参考价)"
        r"(?:合计|总计|总额|为|是|：|:|\s)*"
        r"(?:人民币)?\s*([0-9][0-9,，]*(?:\.[0-9]+)?\s*(?:万元|万|元)?)"
    ),
}
FINANCE_KEYWORDS = ["金融服务", "一键贷款", "最高可贷", "参考利率", "贷款期限", "月供", "可贷比例"]
NOT_APPLICABLE_DEFAULTS = {
    "时间桶开始": "不适用",
    "时间桶结束": "不适用",
    "是否支持保险": "未知",
    "是否配资服务": "未知",
    "purchaseRestriction": "未知",
    "详情接口错误": "无",
    "成交确认书链接": "无",
    "金融机构": "未知",
    "最高可贷比例": "未知",
    "参考利率": "未知",
    "金融其他费用": "未知",
}


def clean_header(value: Any) -> str:
    return str(value or "").strip()


def read_headers(path: Path) -> List[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        return [clean_header(value) for value in row]
    finally:
        workbook.close()


def first_non_empty(row: Dict[str, Any], candidates: Iterable[str]) -> Any:
    for name in candidates:
        value = row.get(name, "")
        if value is None:
            continue
        if str(value).strip() == "":
            continue
        return value
    return ""


def text_blob(row: Dict[str, Any]) -> str:
    parts = []
    for field in ("附件正文文本", "标的物介绍文本", "竞买公告", "拍卖公告", "标的物详情描述"):
        value = row.get(field, "")
        if value is not None and str(value).strip():
            parts.append(str(value))
    return "\n".join(parts)


def parse_money_yuan(value: str) -> Optional[float]:
    text = str(value or "").strip().replace(",", "").replace("，", "")
    if not re.search(r"\d", text):
        return None
    match = re.search(r"([0-9]+(?:\.[0-9]+)?)", text)
    if not match:
        return None
    number = float(match.group(1))
    if "万" in text:
        number *= 10000
    if number <= 0:
        return None
    return round(number, 2)


def derived_price_yuan(row: Dict[str, Any], kind: str) -> Any:
    blob = text_blob(row)
    if not blob:
        return ""
    pattern = PRICE_PATTERNS.get(kind)
    if not pattern:
        return ""
    for match in pattern.finditer(blob):
        value = parse_money_yuan(match.group(1))
        if value is not None:
            return int(value) if float(value).is_integer() else value
    return ""


def attachment_download_links(row: Dict[str, Any]) -> str:
    existing = first_non_empty(row, ("附件链接",))
    if existing:
        return existing
    raw_ids = str(row.get("附件ID", "") or "")
    ids = [part.strip() for part in re.split(r"[；;|,\n]+", raw_ids) if part.strip()]
    if not ids:
        return ""
    return "；".join(f"https://sf.taobao.com/download_attach.htm?attach_id={attach_id}" for attach_id in ids)


def finance_snippet_from_text(row: Dict[str, Any]) -> str:
    blob = text_blob(row)
    if not blob:
        return ""
    hits = []
    for keyword in FINANCE_KEYWORDS:
        index = blob.find(keyword)
        if index < 0:
            continue
        start = max(0, index - 80)
        end = min(len(blob), index + 220)
        snippet = re.sub(r"\s+", " ", blob[start:end]).strip()
        if snippet and snippet not in hits:
            hits.append(snippet)
        if len(hits) >= 2:
            break
    return "；".join(hits)


def finance_detail_value(target: str, row: Dict[str, Any]) -> str:
    direct = first_non_empty(row, (target,))
    if direct:
        return str(direct)
    snippet = finance_snippet_from_text(row)
    if not snippet:
        return "未知"
    if target == "最高可贷比例":
        match = re.search(r"(?:最高可贷|可贷比例|贷款成数)[^0-9一二三四五六七八九十]{0,20}([0-9一二三四五六七八九十]{1,3}\s*(?:成|%|％))", snippet)
        if match:
            return match.group(1).replace(" ", "")
    if target == "参考利率":
        match = re.search(r"(?:参考利率|利率)[^0-9]{0,20}([0-9]+(?:\.[0-9]+)?\s*[%％])", snippet)
        if match:
            return match.group(1).replace(" ", "")
    if target == "金融机构":
        match = re.search(r"([\u4e00-\u9fff]{2,20}(?:银行|金融|担保|贷款)[\u4e00-\u9fff]{0,12})", snippet)
        if match:
            return match.group(1)
    if target == "金融其他费用":
        return snippet[:300]
    return "未知"


def normalize_id(value: Any) -> str:
    text = str(value or "").strip()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text if text.isdigit() else text


def read_jsonl_by_id(path: Path) -> Dict[str, Dict[str, Any]]:
    result: Dict[str, Dict[str, Any]] = {}
    if not path.exists():
        return result
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if not line:
                continue
            try:
                row = json.loads(line)
            except Exception:
                continue
            if not isinstance(row, dict):
                continue
            item_id = normalize_id(row.get("标的物ID"))
            if item_id:
                result[item_id] = row
    return result


def attachment_status(row: Dict[str, Any]) -> str:
    explicit = first_non_empty(row, ("附件抓取状态", "附件正文抓取状态"))
    if explicit:
        return explicit
    count = first_non_empty(row, ("附件数量",))
    names = first_non_empty(row, ("附件名称", "附件链接", "附件索引原文"))
    if str(count).strip() not in ("", "0") or names:
        return "成功"
    return ""


def attachment_time(row: Dict[str, Any]) -> Any:
    if not attachment_status(row):
        return ""
    return first_non_empty(row, ("附件抓取时间", "附件正文抓取时间", "详情抓取时间"))


def finance_text(row: Dict[str, Any]) -> str:
    parts = []
    for field in ("金融机构", "最高可贷比例", "参考利率", "金融其他费用"):
        value = first_non_empty(row, (field,))
        if value:
            parts.append(f"{field}:{value}")
    return "；".join(parts)


def mapped_value(target: str, row: Dict[str, Any]) -> Any:
    if target == "skuId":
        return first_non_empty(row, ("skuId", "标的物ID"))
    if target == "shopId":
        return first_non_empty(row, ("shopId", "处置机构", "处置法院", "法院/处置机构", "法院")) or "阿里法拍"
    if target == "vendorId":
        return first_non_empty(row, ("vendorId", "处置机构", "处置法院", "法院/处置机构", "法院")) or "阿里法拍"
    if target in {"金融机构", "最高可贷比例", "参考利率", "金融其他费用"}:
        return finance_detail_value(target, row)
    if target in NOT_APPLICABLE_DEFAULTS:
        return first_non_empty(row, (target,)) or NOT_APPLICABLE_DEFAULTS[target]
    if target == "附件抓取状态":
        return attachment_status(row)
    if target == "附件抓取时间":
        return attachment_time(row)
    if target == "附件链接":
        return attachment_download_links(row)
    if target == "金融服务原文":
        return first_non_empty(row, ALIASES[target]) or finance_text(row) or finance_snippet_from_text(row)
    if target in row and str(row.get(target, "")).strip() != "":
        return row[target]
    aliases = ALIASES.get(target)
    if aliases:
        value = first_non_empty(row, aliases)
        if value != "":
            return value
    if target in {"评估价", "评估价_元", "评估价_详情_元"}:
        return derived_price_yuan(row, "评估价")
    if target in {"市场价", "市场价_元", "市场价_详情_元"}:
        return derived_price_yuan(row, "市场价")
    return ""


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Align Ali Guangdong output to JD final table schema.")
    parser.add_argument("--ali-input", default=str(DEFAULT_ALI_INPUT))
    parser.add_argument("--ali-index", default=str(DEFAULT_ALI_INDEX))
    parser.add_argument("--ali-detail-jsonl", default=str(DEFAULT_ALI_DETAIL_JSONL))
    parser.add_argument("--jd-schema", default=str(DEFAULT_JD_SCHEMA))
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT))
    parser.add_argument("--report", default=str(DEFAULT_REPORT))
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    ali_input = Path(args.ali_input)
    ali_index = Path(args.ali_index)
    ali_detail_jsonl = Path(args.ali_detail_jsonl)
    jd_schema = Path(args.jd_schema)
    output = Path(args.output)
    report = Path(args.report)

    target_headers = read_headers(jd_schema)
    output.parent.mkdir(parents=True, exist_ok=True)

    detail_rows = read_jsonl_by_id(ali_detail_jsonl)
    source_excel = ali_index if ali_index.exists() and detail_rows else ali_input
    ali_workbook = load_workbook(source_excel, read_only=True, data_only=True)
    try:
        ali_sheet = ali_workbook.active
        source_header_row = next(ali_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        source_headers = [clean_header(value) for value in source_header_row]

        tmp_path = output.with_suffix(output.suffix + ".tmp.xlsx")
        workbook = xlsxwriter.Workbook(
            str(tmp_path),
            {"strings_to_urls": False, "constant_memory": True, "use_zip64": True},
        )
        worksheet = workbook.add_worksheet("Sheet1")
        for col_index, header in enumerate(target_headers):
            worksheet.write(0, col_index, header)

        row_count = 0
        non_empty_by_column = {header: 0 for header in target_headers}
        for excel_row_index, values in enumerate(ali_sheet.iter_rows(min_row=2, values_only=True), start=1):
            source_row = {
                source_headers[index]: values[index] if index < len(values) and values[index] is not None else ""
                for index in range(len(source_headers))
            }
            item_id = normalize_id(first_non_empty(source_row, ("标的物ID", "拍品ID", "itemId", "item_id", "id")))
            detail_row = detail_rows.get(item_id, {})
            if detail_row:
                merged_row = dict(source_row)
                merged_row.update(detail_row)
            else:
                merged_row = source_row
            output_values = [mapped_value(header, merged_row) for header in target_headers]
            for col_index, value in enumerate(output_values):
                worksheet.write(excel_row_index, col_index, value)
                if str(value or "").strip() != "":
                    non_empty_by_column[target_headers[col_index]] += 1
            row_count += 1
            if row_count % 50000 == 0:
                print(f"align_progress rows={row_count}", flush=True)
        workbook.close()
        tmp_path.replace(output)
    finally:
        ali_workbook.close()

    output_headers = read_headers(output)
    report_payload = {
        "verified_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "ali_input": str(ali_input),
        "ali_source_excel": str(source_excel),
        "ali_index": str(ali_index),
        "ali_detail_jsonl": str(ali_detail_jsonl),
        "ali_detail_jsonl_rows_loaded": len(detail_rows),
        "jd_schema": str(jd_schema),
        "output": str(output),
        "rows": row_count,
        "columns": len(output_headers),
        "headers_match_jd": output_headers == target_headers,
        "target_headers": target_headers,
        "output_headers": output_headers,
        "non_empty_by_column": non_empty_by_column,
    }
    report.write_text(json.dumps(report_payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(
        json.dumps(
            {
                "rows": row_count,
                "columns": len(output_headers),
                "headers_match_jd": output_headers == target_headers,
                "output": str(output),
                "report": str(report),
            },
            ensure_ascii=False,
            indent=2,
        ),
        flush=True,
    )


if __name__ == "__main__":
    main()
