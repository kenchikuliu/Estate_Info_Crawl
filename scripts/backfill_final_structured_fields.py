# -*- coding: utf-8 -*-
"""Backfill structured estate fields from existing long-text columns.

This is a post-processing repair step for already merged final workbooks. It
does not refetch network data; it only extracts fields already present in
description/notice/right-status text and appends traceability columns.
"""
from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Sequence

import xlsxwriter
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from utils.jd_detail_parser import (  # noqa: E402
    clean_text,
    extract_labeled_fields,
    extract_rights_status_text,
    postprocess_structured_fields,
    sanitize_structured_value,
)


TARGET_FIELDS = ["建筑面积", "竣工时间", "所在层", "总层数"]
SOURCE_TEXT_FIELDS = [
    "标的物介绍文本",
    "权利限制状况及抵押状况",
    "房屋权属状况",
    "土地权属状况",
    "标的物详情描述",
    "竞买公告文本",
    "竞买公告",
    "拍卖公告",
    "附件索引原文",
]
TRACE_COLUMNS = ["结构字段补全来源", "结构字段补全项", "结构字段补全时间"]


def configure_output_encoding() -> None:
    for stream_name in ("stdout", "stderr"):
        stream = getattr(sys, stream_name, None)
        if hasattr(stream, "reconfigure"):
            try:
                stream.reconfigure(encoding="utf-8", errors="backslashreplace")
            except Exception:
                pass


def is_empty(value: Any) -> bool:
    text = clean_text(str(value or ""))
    return not text or text.lower() in {"nan", "none", "null"}


def normalized_existing_value(field: str, value: Any) -> str:
    return sanitize_structured_value(field, str(value or ""))


def unique_headers(raw_headers: Sequence[Any]) -> List[str]:
    result: List[str] = []
    seen: Dict[str, int] = {}
    for index, value in enumerate(raw_headers):
        base = clean_text(str(value or "")) or f"列{index + 1}"
        count = seen.get(base, 0)
        seen[base] = count + 1
        result.append(base if count == 0 else f"{base}_{count + 1}")
    return result


def text_blob(row: Dict[str, Any]) -> str:
    parts: List[str] = []
    for field in SOURCE_TEXT_FIELDS:
        value = clean_text(str(row.get(field, "") or ""))
        if value and value not in parts:
            parts.append(value)
    return "\n".join(parts)


def parse_structured_fields(row: Dict[str, Any]) -> Dict[str, str]:
    blob = text_blob(row)
    if not blob:
        return {}
    parsed: Dict[str, str] = {}
    parsed.update(extract_labeled_fields(blob))
    parsed.update(extract_rights_status_text(blob))
    return postprocess_structured_fields(parsed)


def write_empty_workbook(path: Path, headers: Sequence[str]) -> None:
    workbook = xlsxwriter.Workbook(str(path), {"strings_to_urls": False, "constant_memory": True, "use_zip64": True})
    worksheet = workbook.add_worksheet("Sheet1")
    for col_index, header in enumerate(headers):
        worksheet.write(0, col_index, header)
    workbook.close()


def backfill_workbook(input_path: Path, output_path: Path, report_path: Path) -> Dict[str, Any]:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = output_path.with_suffix(output_path.suffix + ".tmp.xlsx")

    workbook_in = load_workbook(input_path, read_only=True, data_only=True)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    try:
        sheet = workbook_in.active
        rows_iter = sheet.iter_rows(values_only=True)
        raw_headers = next(rows_iter, None)
        if not raw_headers:
            write_empty_workbook(tmp_path, [])
            tmp_path.replace(output_path)
            report = {"created_at": now, "input": str(input_path), "output": str(output_path), "rows": 0}
            report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
            return report

        headers = unique_headers(raw_headers)
        output_headers = list(headers)
        for field in TARGET_FIELDS:
            if field not in output_headers:
                output_headers.append(field)
        for col in TRACE_COLUMNS:
            if col not in output_headers:
                output_headers.append(col)

        workbook_out = xlsxwriter.Workbook(
            str(tmp_path),
            {"strings_to_urls": False, "constant_memory": True, "use_zip64": True},
        )
        worksheet = workbook_out.add_worksheet("Sheet1")
        for col_index, header in enumerate(output_headers):
            worksheet.write(0, col_index, header)

        total = 0
        filled_counts = {field: 0 for field in TARGET_FIELDS}
        repaired_counts = {field: 0 for field in TARGET_FIELDS}
        pre_existing_counts = {field: 0 for field in TARGET_FIELDS}
        sample_rows: List[Dict[str, Any]] = []

        for output_row_index, values in enumerate(rows_iter, start=1):
            total += 1
            row = {
                headers[index]: values[index] if index < len(values) and values[index] is not None else ""
                for index in range(len(headers))
            }
            parsed = parse_structured_fields(row)
            filled_fields: List[str] = []
            repaired_fields: List[str] = []
            for field in TARGET_FIELDS:
                original_value = row.get(field, "")
                current_value = normalized_existing_value(field, original_value)
                if current_value:
                    pre_existing_counts[field] += 1
                    if clean_text(str(original_value or "")) != current_value:
                        row[field] = current_value
                        repaired_counts[field] += 1
                        repaired_fields.append(field)
                    continue
                value = parsed.get(field, "")
                if value:
                    row[field] = value
                    if is_empty(original_value):
                        filled_counts[field] += 1
                        filled_fields.append(field)
                    else:
                        repaired_counts[field] += 1
                        repaired_fields.append(field)

            changed_fields = filled_fields + repaired_fields
            if changed_fields:
                row["结构字段补全来源"] = "已有长文本字段"
                row["结构字段补全项"] = "；".join(changed_fields)
                row["结构字段补全时间"] = now
                if len(sample_rows) < 20:
                    sample_rows.append(
                        {
                            "row": total + 1,
                            "id": clean_text(str(row.get("标的物ID", ""))),
                            "filled": {field: row.get(field, "") for field in filled_fields},
                            "repaired": {field: row.get(field, "") for field in repaired_fields},
                        }
                    )
            elif "结构字段补全来源" not in row:
                row["结构字段补全来源"] = ""
                row["结构字段补全项"] = ""
                row["结构字段补全时间"] = ""

            for col_index, header in enumerate(output_headers):
                worksheet.write(output_row_index, col_index, row.get(header, ""))

        workbook_out.close()
        tmp_path.replace(output_path)
    finally:
        workbook_in.close()

    report = {
        "created_at": now,
        "input": str(input_path),
        "output": str(output_path),
        "rows": total,
        "target_fields": TARGET_FIELDS,
        "source_text_fields": SOURCE_TEXT_FIELDS,
        "pre_existing_counts": pre_existing_counts,
        "filled_counts": filled_counts,
        "repaired_counts": repaired_counts,
        "samples": sample_rows,
    }
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    return report


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--report", required=True)
    return parser.parse_args()


def main() -> None:
    configure_output_encoding()
    args = parse_args()
    report = backfill_workbook(Path(args.input), Path(args.output), Path(args.report))
    print(json.dumps(report, ensure_ascii=False, indent=2), flush=True)


if __name__ == "__main__":
    main()
