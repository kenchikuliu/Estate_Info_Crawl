# -*- coding: utf-8 -*-
"""
Supervise the long Alibaba API detail backfill until JSONL and Excel are ready.

It intentionally launches the fetcher as child processes instead of importing it,
so an interrupted batch can be resumed from the JSONL checkpoint cleanly.
"""
from __future__ import annotations

import argparse
import json
import subprocess
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INPUT = PROJECT_ROOT / r"output\ali_sf_gd_full_h5_20260611\阿里法拍房_广东_全量索引.xlsx"
DEFAULT_JSONL = PROJECT_ROOT / r"output\ali_sf_gd_full_h5_20260611\阿里法拍房_广东_详情回填_API.jsonl"
DEFAULT_ATTACHMENT_JSONL = PROJECT_ROOT / r"output\ali_sf_gd_full_h5_20260611\阿里法拍房_广东_附件正文回填_API.jsonl"
DEFAULT_COMBINED_JSONL = PROJECT_ROOT / r"output\ali_sf_gd_full_h5_20260611\阿里法拍房_广东_详情回填_API_含附件正文.jsonl"
DEFAULT_OUTPUT = PROJECT_ROOT / r"output\ali_sf_gd_full_h5_20260611\阿里法拍房_广东_详情回填_API.xlsx"
DEFAULT_CHECKPOINT = PROJECT_ROOT / r"output\ali_sf_gd_full_h5_20260611\阿里法拍房_广东_详情回填_API.checkpoint.json"
DEFAULT_REPORT = PROJECT_ROOT / r"output\ali_sf_gd_full_h5_20260611\阿里法拍房_广东_详情回填_API.verify.json"


def count_excel_data_rows(path: Path) -> int:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        return max(0, (sheet.max_row or 1) - 1)
    finally:
        workbook.close()


def read_jsonl_rows(path: Path) -> List[Dict[str, Any]]:
    if not path.exists():
        return []
    rows: List[Dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if not line:
                continue
            try:
                item = json.loads(line)
            except Exception:
                continue
            if isinstance(item, dict):
                rows.append(item)
    return rows


def coverage(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    fields = [
        "详情抓取状态",
        "标的物介绍文本",
        "附件名称",
        "建筑面积",
        "权证情况",
        "房屋用途",
        "所在层",
        "成交价/获拍价_元",
    ]
    ids = [str(row.get("标的物ID", "")).strip() for row in rows if str(row.get("标的物ID", "")).strip()]
    result: Dict[str, Any] = {
        "rows": len(rows),
        "unique_ids": len(set(ids)),
        "duplicate_rows": len(ids) - len(set(ids)),
    }
    for field in fields:
        non_empty = sum(1 for row in rows if str(row.get(field, "")).strip() not in ("", "0"))
        result[field] = {"non_empty": non_empty, "rate": round(non_empty / len(rows), 4) if rows else 0}
    status_counts: Dict[str, int] = {}
    for row in rows:
        status = str(row.get("详情抓取状态") or "")
        status_counts[status] = status_counts.get(status, 0) + 1
    result["status_counts"] = status_counts
    return result


def run_command(command: List[str], log_path: Path) -> int:
    log_path.parent.mkdir(parents=True, exist_ok=True)
    with log_path.open("a", encoding="utf-8") as log:
        log.write(f"\n[{datetime.now():%Y-%m-%d %H:%M:%S}] run {' '.join(command)}\n")
        log.flush()
        process = subprocess.Popen(
            command,
            cwd=str(PROJECT_ROOT),
            stdout=log,
            stderr=subprocess.STDOUT,
            text=True,
        )
        return process.wait()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Supervise Ali API detail backfill to completion.")
    parser.add_argument("--input", default=str(DEFAULT_INPUT))
    parser.add_argument("--jsonl-output", default=str(DEFAULT_JSONL))
    parser.add_argument("--attachment-jsonl", default=str(DEFAULT_ATTACHMENT_JSONL))
    parser.add_argument("--combined-jsonl", default=str(DEFAULT_COMBINED_JSONL))
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT))
    parser.add_argument("--checkpoint", default=str(DEFAULT_CHECKPOINT))
    parser.add_argument("--report", default=str(DEFAULT_REPORT))
    parser.add_argument("--workers", type=int, default=6)
    parser.add_argument("--save-every", type=int, default=500)
    parser.add_argument("--timeout", type=int, default=25)
    parser.add_argument("--retries", type=int, default=3)
    parser.add_argument("--raw-text-limit", type=int, default=2000)
    parser.add_argument("--poll-seconds", type=int, default=300)
    parser.add_argument("--max-fetch-runs", type=int, default=5)
    parser.add_argument("--wait-existing-pid", type=int, default=0, help="Wait for an already-running fetch process before launching resume runs.")
    parser.add_argument("--skip-attachment-text", action="store_true")
    parser.add_argument("--attachment-workers", type=int, default=3)
    parser.add_argument("--attachment-limit", type=int, default=2)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_path = Path(args.input)
    jsonl_path = Path(args.jsonl_output)
    attachment_jsonl_path = Path(args.attachment_jsonl)
    combined_jsonl_path = Path(args.combined_jsonl)
    output_path = Path(args.output)
    checkpoint_path = Path(args.checkpoint)
    report_path = Path(args.report)
    log_path = PROJECT_ROOT / "logs" / f"ali_sf_api_detail_supervise_{datetime.now():%Y%m%d_%H%M%S}.log"

    target_rows = count_excel_data_rows(input_path)
    fetch_runs = 0
    if args.wait_existing_pid > 0:
        while True:
            try:
                import psutil  # type: ignore

                running = psutil.pid_exists(args.wait_existing_pid)
            except Exception:
                running = subprocess.run(
                    ["powershell", "-NoProfile", "-Command", f"Get-Process -Id {args.wait_existing_pid} -ErrorAction SilentlyContinue"],
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.DEVNULL,
                    check=False,
                ).returncode == 0
            if not running:
                break
            print(f"waiting_existing_pid pid={args.wait_existing_pid}", flush=True)
            time.sleep(args.poll_seconds)

    while True:
        rows = read_jsonl_rows(jsonl_path)
        ids = [str(row.get("标的物ID", "")).strip() for row in rows if str(row.get("标的物ID", "")).strip()]
        done = len(set(ids))
        duplicate_rows = len(ids) - done
        print(
            f"supervise_status target={target_rows} jsonl_rows={len(rows)} "
            f"unique_ids={done} duplicate_rows={duplicate_rows} fetch_runs={fetch_runs}",
            flush=True,
        )
        if done >= target_rows:
            break
        if fetch_runs >= args.max_fetch_runs:
            raise RuntimeError(f"fetch runs exhausted: unique_ids={done} target={target_rows}")
        command = [
            sys.executable,
            "scripts/ali_sf_api_detail_backfill.py",
            "--input",
            str(input_path),
            "--output",
            str(output_path),
            "--jsonl-output",
            str(jsonl_path),
            "--checkpoint",
            str(checkpoint_path),
            "--workers",
            str(args.workers),
            "--save-every",
            str(args.save_every),
            "--timeout",
            str(args.timeout),
            "--retries",
            str(args.retries),
            "--raw-text-limit",
            str(args.raw_text_limit),
            "--no-export",
        ]
        fetch_runs += 1
        code = run_command(command, log_path)
        if code != 0:
            print(f"fetch_run_exit code={code}; sleeping before retry", flush=True)
            time.sleep(args.poll_seconds)

    export_jsonl_path = jsonl_path
    if not args.skip_attachment_text:
        attachment_code = run_command(
            [
                sys.executable,
                "scripts/ali_sf_attachment_text_backfill.py",
                "--detail-jsonl",
                str(jsonl_path),
                "--output-jsonl",
                str(attachment_jsonl_path),
                "--combined-jsonl",
                str(combined_jsonl_path),
                "--workers",
                str(args.attachment_workers),
                "--attachment-limit",
                str(args.attachment_limit),
                "--empty-desc-only",
                "--timeout",
                "45",
                "--raw-text-limit",
                str(max(args.raw_text_limit, 4000)),
            ],
            log_path,
        )
        if attachment_code != 0:
            raise RuntimeError(f"attachment text backfill failed code={attachment_code}")
        export_jsonl_path = combined_jsonl_path

    export_code = run_command(
        [
            sys.executable,
            "scripts/ali_sf_api_detail_backfill.py",
            "--input",
            str(input_path),
            "--output",
            str(output_path),
            "--jsonl-output",
            str(export_jsonl_path),
            "--checkpoint",
            str(checkpoint_path),
            "--export-only",
            "--include-unfinished-index",
        ],
        log_path,
    )
    if export_code != 0:
        raise RuntimeError(f"export failed code={export_code}")

    rows = read_jsonl_rows(export_jsonl_path)
    coverage_result = coverage(rows)
    report = {
        "verified_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "input_rows": target_rows,
        "jsonl_rows": len(rows),
        "jsonl_unique_ids": coverage_result["unique_ids"],
        "jsonl_duplicate_rows": coverage_result["duplicate_rows"],
        "excel_output": str(output_path),
        "jsonl_output": str(export_jsonl_path),
        "base_jsonl_output": str(jsonl_path),
        "attachment_jsonl_output": str(attachment_jsonl_path),
        "log": str(log_path),
        "coverage": coverage_result,
    }
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"supervise_done report={report_path} output={output_path}", flush=True)


if __name__ == "__main__":
    main()
