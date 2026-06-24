# -*- coding: utf-8 -*-
"""
Merge JD and Alibaba judicial-auction final tables for one region.

Both inputs are expected to already use the same JD-compatible column schema.
The output preserves that schema and writes a JSON report with the dedupe
counts. Cross-platform dedupe is intentionally conservative to avoid merging
different auction rounds for the same property.
"""
from __future__ import annotations

import argparse
import html
import json
import re
import sys
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import xlsxwriter
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]

MISSING = {"", "none", "null", "nan"}


def configure_output_encoding() -> None:
    for stream_name in ("stdout", "stderr"):
        stream = getattr(sys, stream_name, None)
        if hasattr(stream, "reconfigure"):
            try:
                stream.reconfigure(encoding="utf-8", errors="backslashreplace")
            except Exception:
                pass


def print_report(report: Dict[str, Any]) -> None:
    text = json.dumps(report, ensure_ascii=False, indent=2)
    try:
        print(text, flush=True)
    except UnicodeEncodeError:
        print(text.encode("utf-8", errors="backslashreplace").decode("utf-8"), flush=True)

ID_COL = "标的物ID"
LINK_COL = "链接"
TITLE_COL = "标题"
ITEM_NAME_COL = "标的物名称"
TARGET_NAME_COL = "标的名称"
ADDR_COL = "地址"
DETAIL_ADDR_COL = "详情地址"
AREA_COL = "建筑面积"
CURRENT_PRICE_COL = "当前价_元"
START_PRICE_COL = "起拍价_元"
ASSESS_PRICE_COL = "评估价_元"
MARKET_PRICE_COL = "市场价_元"
END_TIME_COL = "结束时间"
DEAL_TIME_COL = "成交时间"
COURT_COL = "处置机构"
DETAIL_COURT_COL = "法院/处置机构"
PUBLISH_SOURCE_COL = "publishSource"
ATTACH_STATUS_COL = "附件抓取状态"
ATTACH_COUNT_COL = "附件数量"
INTRO_COL = "标的物介绍文本"
NOTICE_COL = "竞买公告文本"
ATTACH_INDEX_COL = "附件索引原文"

SOURCE_LABELS = {
    "jd": "京东法拍",
    "ali": "阿里法拍",
}


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in MISSING:
        return ""
    return text


def normalize_id(value: Any) -> str:
    text = clean_text(value)
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text


def normalize_url(value: Any) -> str:
    text = clean_text(value)
    text = text.replace("\\", "/").strip()
    return text.rstrip("/")


def normalize_text_key(value: Any) -> str:
    text = html.unescape(clean_text(value))
    text = unicodedata.normalize("NFKC", text).lower()
    replacements = [
        "拍卖公告",
        "第一次拍卖",
        "第二次拍卖",
        "一拍",
        "二拍",
        "变卖",
        "司法拍卖",
        "网络司法拍卖",
        "房产一套",
        "房屋一套",
        "房地产",
        "房产",
        "房屋",
        "住宅",
        "位于",
        "标的物",
    ]
    for token in replacements:
        text = text.replace(token, "")
    text = re.sub(r"[\\/:*?\"<>|,，。；;、\s\-_—（）()【】\[\]{}]+", "", text)
    return text


def parse_money(value: Any) -> str:
    text = clean_text(value).replace(",", "").replace("，", "")
    if not text:
        return ""
    match = re.search(r"([0-9]+(?:\.[0-9]+)?)", text)
    if not match:
        return ""
    number = float(match.group(1))
    if "万" in text:
        number *= 10000
    if number <= 0:
        return ""
    return str(int(round(number)))


def parse_area(value: Any) -> str:
    text = unicodedata.normalize("NFKC", clean_text(value)).replace(",", "")
    if not text:
        return ""
    match = re.search(r"([0-9]+(?:\.[0-9]+)?)", text)
    if not match:
        return ""
    number = round(float(match.group(1)), 2)
    if number <= 0:
        return ""
    return f"{number:.2f}"


def parse_date(value: Any) -> str:
    text = clean_text(value)
    if not text:
        return ""
    match = re.search(r"(20[0-9]{2}|19[0-9]{2})[-/.年]([01]?[0-9])[-/.月]([0-3]?[0-9])", text)
    if not match:
        return text[:10] if len(text) >= 10 else text
    return f"{int(match.group(1)):04d}-{int(match.group(2)):02d}-{int(match.group(3)):02d}"


def read_headers(path: Path) -> List[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        row = next(workbook.active.iter_rows(min_row=1, max_row=1, values_only=True))
        return [clean_text(value) for value in row]
    finally:
        workbook.close()


def row_get(row: Sequence[Any], indexes: Dict[str, int], name: str) -> Any:
    index = indexes.get(name)
    if index is None or index >= len(row):
        return ""
    return row[index]


def first_non_empty(row: Sequence[Any], indexes: Dict[str, int], names: Iterable[str]) -> str:
    for name in names:
        value = clean_text(row_get(row, indexes, name))
        if value:
            return value
    return ""


def meaningful_texts(row: Sequence[Any], indexes: Dict[str, int]) -> List[str]:
    texts = []
    for name in (DETAIL_ADDR_COL, ADDR_COL, ITEM_NAME_COL, TARGET_NAME_COL, TITLE_COL):
        value = normalize_text_key(row_get(row, indexes, name))
        if len(value) >= 8 and value not in texts:
            texts.append(value)
    return texts


def cross_keys(row: Sequence[Any], indexes: Dict[str, int]) -> List[str]:
    texts = meaningful_texts(row, indexes)
    if not texts:
        return []
    area = parse_area(first_non_empty(row, indexes, (AREA_COL,)))
    start = parse_money(first_non_empty(row, indexes, (START_PRICE_COL,)))
    current = parse_money(first_non_empty(row, indexes, (CURRENT_PRICE_COL,)))
    assess = parse_money(first_non_empty(row, indexes, (ASSESS_PRICE_COL, MARKET_PRICE_COL)))
    date = parse_date(first_non_empty(row, indexes, (END_TIME_COL, DEAL_TIME_COL)))
    court = normalize_text_key(first_non_empty(row, indexes, (DETAIL_COURT_COL, COURT_COL)))
    keys: List[str] = []
    for text in texts:
        if area and start and current and date:
            keys.append(f"tacd|{text}|{area}|{start}|{current}|{date}")
        if area and start and date:
            keys.append(f"tad|{text}|{area}|{start}|{date}")
        if area and assess and date:
            keys.append(f"taed|{text}|{area}|{assess}|{date}")
        if start and current and date and court:
            keys.append(f"tpcd|{text}|{start}|{current}|{date}|{court}")
        if area and start and current:
            keys.append(f"tac|{text}|{area}|{start}|{current}")
    return keys


def platform_key(source: str, row: Sequence[Any], indexes: Dict[str, int]) -> str:
    link = normalize_url(row_get(row, indexes, LINK_COL))
    if link:
        return f"{source}|link|{link}"
    item_id = normalize_id(row_get(row, indexes, ID_COL))
    if item_id:
        return f"{source}|id|{item_id}"
    title = normalize_text_key(row_get(row, indexes, TITLE_COL))
    return f"{source}|row|{title}|{parse_date(row_get(row, indexes, END_TIME_COL))}"


def non_empty_count(row: Sequence[Any]) -> int:
    return sum(1 for value in row if clean_text(value))


def row_score(row: Sequence[Any], indexes: Dict[str, int]) -> int:
    score = non_empty_count(row)
    for name, weight in (
        (ITEM_NAME_COL, 3),
        (DETAIL_ADDR_COL, 3),
        (AREA_COL, 4),
        (ASSESS_PRICE_COL, 3),
        (MARKET_PRICE_COL, 3),
        (ATTACH_STATUS_COL, 2),
        (INTRO_COL, 6),
        (NOTICE_COL, 4),
        (ATTACH_INDEX_COL, 4),
    ):
        if clean_text(row_get(row, indexes, name)):
            score += weight
    try:
        attach_count = int(float(clean_text(row_get(row, indexes, ATTACH_COUNT_COL)) or "0"))
    except ValueError:
        attach_count = 0
    score += min(attach_count, 10)
    return score


def set_source(row: List[Any], indexes: Dict[str, int], source_label: str) -> None:
    index = indexes.get(PUBLISH_SOURCE_COL)
    if index is None:
        return
    current = clean_text(row[index])
    parts = [part for part in re.split(r"[；;|,，\s]+", current) if part]
    if source_label not in parts:
        parts.append(source_label)
    row[index] = "；".join(parts) if parts else source_label


def merge_values(base: List[Any], incoming: Sequence[Any], indexes: Dict[str, int]) -> List[Any]:
    merged = list(base)
    for index, value in enumerate(incoming):
        if index >= len(merged):
            continue
        if not clean_text(merged[index]) and clean_text(value):
            merged[index] = value
    return merged


class DedupeStore:
    def __init__(self, headers: List[str]):
        self.headers = headers
        self.indexes = {header: index for index, header in enumerate(headers)}
        self.rows: List[List[Any]] = []
        self.sources: List[set[str]] = []
        self.scores: List[int] = []
        self.platform_map: Dict[str, int] = {}
        self.cross_map: Dict[str, int] = {}
        self.input_rows = 0
        self.platform_duplicates = 0
        self.cross_duplicates = 0
        self.examples: List[Dict[str, Any]] = []

    def add(self, row_values: Sequence[Any], source: str) -> None:
        self.input_rows += 1
        row = list(row_values[: len(self.headers)])
        if len(row) < len(self.headers):
            row.extend([""] * (len(self.headers) - len(row)))
        source_label = SOURCE_LABELS[source]
        set_source(row, self.indexes, source_label)
        score = row_score(row, self.indexes)

        pkey = platform_key(source, row, self.indexes)
        existing_index = self.platform_map.get(pkey)
        duplicate_type = "platform" if existing_index is not None else ""

        if existing_index is None:
            for key in cross_keys(row, self.indexes):
                candidate = self.cross_map.get(key)
                if candidate is not None and source not in self.sources[candidate]:
                    existing_index = candidate
                    duplicate_type = "cross"
                    break

        if existing_index is None:
            index = len(self.rows)
            self.rows.append(row)
            self.sources.append({source})
            self.scores.append(score)
            self.platform_map[pkey] = index
            for key in cross_keys(row, self.indexes):
                self.cross_map.setdefault(key, index)
            return

        if duplicate_type == "platform":
            self.platform_duplicates += 1
        else:
            self.cross_duplicates += 1
            if len(self.examples) < 20:
                self.examples.append(
                    {
                        "kept_id": clean_text(row_get(self.rows[existing_index], self.indexes, ID_COL)),
                        "incoming_id": clean_text(row_get(row, self.indexes, ID_COL)),
                        "kept_title": clean_text(row_get(self.rows[existing_index], self.indexes, TITLE_COL)),
                        "incoming_title": clean_text(row_get(row, self.indexes, TITLE_COL)),
                        "sources": sorted(self.sources[existing_index] | {source}),
                    }
                )

        existing = self.rows[existing_index]
        if score > self.scores[existing_index]:
            merged = merge_values(row, existing, self.indexes)
            self.scores[existing_index] = row_score(merged, self.indexes)
        else:
            merged = merge_values(existing, row, self.indexes)
            self.scores[existing_index] = max(self.scores[existing_index], row_score(merged, self.indexes))
        self.rows[existing_index] = merged
        self.sources[existing_index].add(source)
        set_source(self.rows[existing_index], self.indexes, source_label)
        self.platform_map[pkey] = existing_index
        for key in cross_keys(self.rows[existing_index], self.indexes):
            self.cross_map.setdefault(key, existing_index)


def iter_excel_rows(path: Path, expected_headers: List[str]) -> Iterable[List[Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [clean_text(value) for value in header_row]
        if headers != expected_headers:
            raise ValueError(f"header mismatch for {path}")
        for values in sheet.iter_rows(min_row=2, values_only=True):
            yield list(values)
    finally:
        workbook.close()


def write_excel(path: Path, headers: List[str], rows: Sequence[Sequence[Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_suffix(path.suffix + ".tmp.xlsx")
    workbook = xlsxwriter.Workbook(
        str(tmp_path),
        {"strings_to_urls": False, "constant_memory": True, "use_zip64": True},
    )
    worksheet = workbook.add_worksheet("Sheet1")
    for col, header in enumerate(headers):
        worksheet.write(0, col, header)
    for row_index, row in enumerate(rows, start=1):
        for col, value in enumerate(row):
            worksheet.write(row_index, col, "" if value is None else value)
    workbook.close()
    tmp_path.replace(path)


def merge_region(jd_path: Path, ali_path: Path, output_path: Path, report_path: Path) -> Dict[str, Any]:
    jd_headers = read_headers(jd_path)
    ali_headers = read_headers(ali_path)
    if jd_headers != ali_headers:
        raise ValueError("JD and Ali headers are not identical")

    store = DedupeStore(jd_headers)
    for row in iter_excel_rows(jd_path, jd_headers):
        store.add(row, "jd")
    jd_rows = store.input_rows
    for row in iter_excel_rows(ali_path, jd_headers):
        store.add(row, "ali")
    ali_rows = store.input_rows - jd_rows

    write_excel(output_path, jd_headers, store.rows)
    report = {
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "jd_input": str(jd_path),
        "ali_input": str(ali_path),
        "output": str(output_path),
        "headers": len(jd_headers),
        "jd_rows": jd_rows,
        "ali_rows": ali_rows,
        "input_rows": store.input_rows,
        "output_rows": len(store.rows),
        "removed_rows": store.input_rows - len(store.rows),
        "platform_duplicate_rows": store.platform_duplicates,
        "cross_platform_duplicate_rows": store.cross_duplicates,
        "cross_platform_examples": store.examples,
    }
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    return report


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Merge JD and Ali final tables for a region.")
    parser.add_argument("--jd", required=True)
    parser.add_argument("--ali", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--report", required=True)
    return parser.parse_args()


def main() -> None:
    configure_output_encoding()
    args = parse_args()
    report = merge_region(
        Path(args.jd),
        Path(args.ali),
        Path(args.output),
        Path(args.report),
    )
    print_report(report)


if __name__ == "__main__":
    main()
