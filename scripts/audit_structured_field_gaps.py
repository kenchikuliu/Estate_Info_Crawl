# -*- coding: utf-8 -*-
"""Audit missing structured estate fields that still have clues in long text.

The script is read-only. It scans an existing final/backfilled workbook and
reports rows where a target field is empty while source text still contains a
likely area, build-year, current-floor, or total-floor signal.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Pattern, Sequence, Tuple

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from scripts.backfill_final_structured_fields import SOURCE_TEXT_FIELDS, TARGET_FIELDS  # noqa: E402
from utils.jd_detail_parser import (  # noqa: E402
    clean_text,
    extract_labeled_fields,
    extract_rights_status_text,
    postprocess_structured_fields,
)


FIELD_SIGNAL_PATTERNS: Dict[str, Pattern[str]] = {
    "建筑面积": re.compile(r"建筑面积|建筑总面积|建筑物面积|房屋建筑面积|证载建筑面积|面积"),
    "竣工时间": re.compile(r"竣工|建成|建成年|建成年月|建成日期|建造|建于|建成年份|建成时间"),
    "所在层": re.compile(
        r"所在楼层|所在层|房屋楼层|房屋所在层|楼层|层次|位于第[0-9一二三四五六七八九十百]+层|第[0-9一二三四五六七八九十百]+层"
    ),
    "总层数": re.compile(
        r"总层数|总楼层|房屋总层数|建筑物总层数|建筑物总高|总高|总[0-9一二三四五六七八九十百]+层|共[0-9一二三四五六七八九十百]+层|/[0-9一二三四五六七八九十百]+(?:层|楼|F)|层数"
    ),
}


def configure_output_encoding() -> None:
    for stream_name in ("stdout", "stderr"):
        stream = getattr(sys, stream_name, None)
        if hasattr(stream, "reconfigure"):
            try:
                stream.reconfigure(encoding="utf-8", errors="backslashreplace")
            except Exception:
                pass


def unique_headers(raw_headers: Sequence[Any]) -> List[str]:
    result: List[str] = []
    seen: Dict[str, int] = {}
    for index, value in enumerate(raw_headers):
        base = clean_text(str(value or "")) or f"列{index + 1}"
        count = seen.get(base, 0)
        seen[base] = count + 1
        result.append(base if count == 0 else f"{base}_{count + 1}")
    return result


def is_empty(value: Any) -> bool:
    text = clean_text(str(value or ""))
    return not text or text.lower() in {"nan", "none", "null", "n/a"}


def excerpt_around(text: str, start: int, end: int, radius: int = 90) -> str:
    left = max(0, start - radius)
    right = min(len(text), end + radius)
    prefix = "..." if left > 0 else ""
    suffix = "..." if right < len(text) else ""
    return prefix + clean_text(text[left:right]) + suffix


def find_signal(row: Dict[str, Any], field: str) -> Optional[Tuple[str, str]]:
    pattern = FIELD_SIGNAL_PATTERNS[field]
    for source_field in SOURCE_TEXT_FIELDS:
        text = clean_text(str(row.get(source_field, "") or ""))
        if not text:
            continue
        match = pattern.search(text)
        if match:
            return source_field, excerpt_around(text, match.start(), match.end())
    return None


def text_blob(row: Dict[str, Any]) -> str:
    parts: List[str] = []
    for field in SOURCE_TEXT_FIELDS:
        value = clean_text(str(row.get(field, "") or ""))
        if value and value not in parts:
            parts.append(value)
    return "\n".join(parts)


def normalize_parsed_keys(parsed: Dict[str, str]) -> Dict[str, str]:
    """Guard against console/codepage-decoded keys in long-running shells."""
    if any(field in parsed for field in TARGET_FIELDS):
        return parsed
    key_aliases = {
        "½¨ÖþÃæ»ý": "建筑面积",
        "¿¢¹¤Ê±¼ä": "竣工时间",
        "ËùÔÚ²ã": "所在层",
        "×Ü²ãÊý": "总层数",
    }
    normalized: Dict[str, str] = {}
    for key, value in parsed.items():
        normalized[key_aliases.get(key, key)] = value
    return normalized


def parse_structured_fields(row: Dict[str, Any]) -> Dict[str, str]:
    blob = text_blob(row)
    if not blob:
        return {}
    parsed: Dict[str, str] = {}
    parsed.update(extract_labeled_fields(blob))
    parsed.update(extract_rights_status_text(blob))
    return normalize_parsed_keys(postprocess_structured_fields(parsed))


def first_non_empty(row: Dict[str, Any], fields: Sequence[str]) -> str:
    for field in fields:
        value = clean_text(str(row.get(field, "") or ""))
        if value:
            return value
    return ""


def audit_workbook(input_path: Path, report_path: Path, limit_rows: int, sample_limit: int) -> Dict[str, Any]:
    report_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    try:
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        raw_headers = next(rows_iter, None)
        if not raw_headers:
            report = {
                "created_at": now,
                "input": str(input_path),
                "rows": 0,
                "target_fields": TARGET_FIELDS,
            }
            report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
            return report

        headers = unique_headers(raw_headers)
        missing_counts = {field: 0 for field in TARGET_FIELDS}
        signal_counts = {field: 0 for field in TARGET_FIELDS}
        parsed_signal_counts = {field: 0 for field in TARGET_FIELDS}
        unparsed_signal_counts = {field: 0 for field in TARGET_FIELDS}
        samples: Dict[str, List[Dict[str, Any]]] = {field: [] for field in TARGET_FIELDS}
        total = 0
        scanned_all_rows = True

        for row_number, values in enumerate(rows_iter, start=2):
            if limit_rows and total >= limit_rows:
                scanned_all_rows = False
                break
            total += 1
            row = {
                headers[index]: values[index] if index < len(values) and values[index] is not None else ""
                for index in range(len(headers))
            }
            parsed = parse_structured_fields(row)
            row_id = first_non_empty(row, ["标的物ID", "拍品ID", "paimaiId", "id"])
            title = first_non_empty(row, ["标题", "详情标题", "标的物名称", "标的名称"])

            for field in TARGET_FIELDS:
                if not is_empty(row.get(field)):
                    continue
                missing_counts[field] += 1
                signal = find_signal(row, field)
                if not signal:
                    continue
                source_field, excerpt = signal
                signal_counts[field] += 1
                parsed_value = clean_text(str(parsed.get(field, "") or ""))
                if parsed_value:
                    parsed_signal_counts[field] += 1
                else:
                    unparsed_signal_counts[field] += 1
                if len(samples[field]) < sample_limit:
                    samples[field].append(
                        {
                            "row": row_number,
                            "id": row_id,
                            "title": title,
                            "source_field": source_field,
                            "parsed_value": parsed_value,
                            "excerpt": excerpt,
                        }
                    )
    finally:
        workbook.close()

    report = {
        "created_at": now,
        "input": str(input_path),
        "rows_scanned": total,
        "scanned_all_rows": scanned_all_rows,
        "target_fields": TARGET_FIELDS,
        "source_text_fields": SOURCE_TEXT_FIELDS,
        "missing_counts": missing_counts,
        "missing_with_text_signal_counts": signal_counts,
        "parsed_signal_counts": parsed_signal_counts,
        "unparsed_signal_counts": unparsed_signal_counts,
        "samples": samples,
    }
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    return report


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", required=True)
    parser.add_argument("--report", required=True)
    parser.add_argument("--limit-rows", type=int, default=0, help="0 means scan all rows.")
    parser.add_argument("--sample-limit", type=int, default=20)
    return parser.parse_args()


def main() -> None:
    configure_output_encoding()
    args = parse_args()
    report = audit_workbook(
        input_path=Path(args.input),
        report_path=Path(args.report),
        limit_rows=args.limit_rows,
        sample_limit=args.sample_limit,
    )
    print(json.dumps(report, ensure_ascii=False, indent=2), flush=True)


if __name__ == "__main__":
    main()
