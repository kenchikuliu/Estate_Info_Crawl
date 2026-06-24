# -*- coding: utf-8 -*-
"""Backfill missing JD coordinates in an existing workbook.

This is a controlled post-processing step. It preserves the input workbook,
uses address geocoding only for rows missing longitude/latitude, and writes a
report plus a geocode cache so large runs can be resumed.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Sequence

import xlsxwriter
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from utils.jd_detail_parser import clean_text, geocode_address  # noqa: E402


TRACE_COLUMNS = ["坐标补全时间", "坐标补全地址", "坐标补全状态"]


def configure_output_encoding() -> None:
    for stream_name in ("stdout", "stderr"):
        stream = getattr(sys, stream_name, None)
        if hasattr(stream, "reconfigure"):
            try:
                stream.reconfigure(encoding="utf-8", errors="backslashreplace")
            except Exception:
                pass


def is_empty(value: Any) -> bool:
    text = clean_text(str(value or ""))
    return not text or text.lower() in {"nan", "none", "null"}


def unique_headers(raw_headers: Sequence[Any]) -> List[str]:
    result: List[str] = []
    seen: Dict[str, int] = {}
    for index, value in enumerate(raw_headers):
        base = clean_text(str(value or "")) or f"列{index + 1}"
        count = seen.get(base, 0)
        seen[base] = count + 1
        result.append(base if count == 0 else f"{base}_{count + 1}")
    return result


def first_non_empty(row: Dict[str, Any], fields: Sequence[str]) -> str:
    for field in fields:
        value = clean_text(str(row.get(field, "") or ""))
        if value:
            return value
    return ""


def normalize_id(value: Any) -> str:
    text = clean_text(str(value or ""))
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    match = re.search(r"paimai\.jd\.com/(\d+)", text)
    if match:
        return match.group(1)
    return text if text.isdigit() else ""


def is_jd_row(row: Dict[str, Any]) -> bool:
    platform = first_non_empty(row, ["平台", "来源平台", "数据来源", "平台来源", "publishSource"])
    link = first_non_empty(row, ["链接", "详情链接", "URL", "url"])
    asset_id = first_non_empty(row, ["标的物ID", "拍品ID", "paimaiId", "id"])
    if "京东" in platform or "paimai.jd.com" in link:
        return True
    if "阿里" in platform or "sf.taobao.com" in link:
        return False
    return bool(normalize_id(asset_id))


def candidate_address(row: Dict[str, Any]) -> str:
    address = first_non_empty(
        row,
        [
            "详情地址",
            "地址",
            "标的物地址",
            "所在位置",
            "标的物名称",
            "标的名称",
            "详情标题",
            "标题",
        ],
    )
    if not address:
        return ""
    province = first_non_empty(row, ["详情省", "省份", "省份筛选", "省"])
    city = first_non_empty(row, ["详情市", "城市", "城市筛选", "市"])
    county = first_non_empty(row, ["详情区县", "区县", "区"])
    prefix_parts: List[str] = []
    for part in [province, city, county]:
        if part and part not in address and part not in prefix_parts:
            prefix_parts.append(part)
    if not prefix_parts:
        return address
    if any(token in address for token in ["省", "市", "自治区", "特别行政区"]):
        return address
    return "".join(prefix_parts) + address


def load_cache(path: Path) -> Dict[str, Dict[str, str]]:
    if not path.exists():
        return {}
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}
    return payload if isinstance(payload, dict) else {}


def save_cache(path: Path, cache: Dict[str, Dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")


def backfill_coordinates(
    input_path: Path,
    output_path: Path,
    report_path: Path,
    cache_path: Path,
    limit: int,
    max_new_geocodes: int,
    cache_only: bool,
    sleep_seconds: float,
    timeout: int,
) -> Dict[str, Any]:
    if not cache_only:
        output_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = output_path.with_suffix(output_path.suffix + ".tmp.xlsx") if not cache_only else None
    cache = load_cache(cache_path)

    workbook_in = load_workbook(input_path, read_only=True, data_only=True)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    try:
        sheet = workbook_in.active
        rows_iter = sheet.iter_rows(values_only=True)
        raw_headers = next(rows_iter, None)
        if not raw_headers:
            if not cache_only and tmp_path is not None:
                workbook_out = xlsxwriter.Workbook(str(tmp_path), {"strings_to_urls": False, "constant_memory": True, "use_zip64": True})
                workbook_out.add_worksheet("Sheet1")
                workbook_out.close()
                tmp_path.replace(output_path)
            report = {
                "created_at": now,
                "input": str(input_path),
                "output": "" if cache_only else str(output_path),
                "cache_only": cache_only,
                "rows": 0,
            }
            report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
            return report

        headers = unique_headers(raw_headers)
        output_headers = list(headers)
        for field in ["经度", "纬度", "经纬度来源"]:
            if field not in output_headers:
                output_headers.append(field)
        for field in TRACE_COLUMNS:
            if field not in output_headers:
                output_headers.append(field)

        workbook_out = None
        worksheet = None
        if not cache_only and tmp_path is not None:
            workbook_out = xlsxwriter.Workbook(
                str(tmp_path),
                {"strings_to_urls": False, "constant_memory": True, "use_zip64": True},
            )
            worksheet = workbook_out.add_worksheet("Sheet1")
            for col_index, header in enumerate(output_headers):
                worksheet.write(0, col_index, header)

        total = jd_rows = missing_coords = candidates = filled = failed = 0
        skipped_by_row_limit = skipped_by_new_geocode_limit = 0
        cache_hits = cache_failures = new_geocode_requests = 0
        scan_complete = True
        samples: List[Dict[str, Any]] = []

        for output_row_index, values in enumerate(rows_iter, start=1):
            total += 1
            row = {
                headers[index]: values[index] if index < len(values) and values[index] is not None else ""
                for index in range(len(headers))
            }
            if is_jd_row(row):
                jd_rows += 1
                if is_empty(row.get("经度")) or is_empty(row.get("纬度")):
                    missing_coords += 1
                    address = candidate_address(row)
                    if address:
                        candidates += 1
                        if limit and candidates > limit:
                            skipped_by_row_limit += 1
                        else:
                            skipped_for_new_limit = False
                            geocoded = cache.get(address)
                            if geocoded is None:
                                skipped_for_new_limit = bool(max_new_geocodes and new_geocode_requests >= max_new_geocodes)
                                if skipped_for_new_limit:
                                    skipped_by_new_geocode_limit += 1
                                    geocoded = {}
                                    if cache_only:
                                        scan_complete = False
                                        break
                                else:
                                    geocoded = geocode_address(address, timeout=timeout)
                                    cache[address] = geocoded
                                    new_geocode_requests += 1
                                    if sleep_seconds > 0:
                                        time.sleep(sleep_seconds)
                            elif isinstance(geocoded, dict) and geocoded.get("经度") and geocoded.get("纬度"):
                                cache_hits += 1
                            else:
                                cache_failures += 1
                            lng = geocoded.get("经度", "") if isinstance(geocoded, dict) else ""
                            lat = geocoded.get("纬度", "") if isinstance(geocoded, dict) else ""
                            if lng and lat:
                                row["经度"] = lng
                                row["纬度"] = lat
                                row["经纬度来源"] = "ArcGIS地址地理编码"
                                row["坐标补全时间"] = now
                                row["坐标补全地址"] = address
                                row["坐标补全状态"] = "成功"
                                filled += 1
                                if len(samples) < 20:
                                    samples.append(
                                        {
                                            "row": total + 1,
                                            "id": normalize_id(first_non_empty(row, ["标的物ID", "拍品ID", "paimaiId", "id"])),
                                            "address": address,
                                            "lng": lng,
                                            "lat": lat,
                                        }
                                    )
                            else:
                                row["坐标补全状态"] = "失败"
                                if not skipped_for_new_limit:
                                    failed += 1

            if worksheet is not None:
                for col_index, header in enumerate(output_headers):
                    worksheet.write(output_row_index, col_index, row.get(header, ""))

        if workbook_out is not None and tmp_path is not None:
            workbook_out.close()
            tmp_path.replace(output_path)
    finally:
        workbook_in.close()

    save_cache(cache_path, cache)
    report = {
        "created_at": now,
        "input": str(input_path),
        "output": "" if cache_only else str(output_path),
        "cache": str(cache_path),
        "cache_only": cache_only,
        "rows": total,
        "scan_complete": scan_complete,
        "jd_rows": jd_rows,
        "jd_missing_coords": missing_coords,
        "geocode_candidates": candidates,
        "filled_coords": filled,
        "failed_geocodes": failed,
        "cache_hits": cache_hits,
        "cache_failures": cache_failures,
        "new_geocode_requests": new_geocode_requests,
        "skipped_by_row_limit": skipped_by_row_limit,
        "skipped_by_new_geocode_limit": skipped_by_new_geocode_limit,
        "samples": samples,
    }
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    return report


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--report", required=True)
    parser.add_argument("--cache", required=True)
    parser.add_argument("--cache-only", action="store_true", help="Populate geocode cache without writing an output workbook.")
    parser.add_argument("--limit", type=int, default=0, help="0 means no limit.")
    parser.add_argument("--max-new-geocodes", type=int, default=0, help="0 means no limit. Cached coordinates are still reused.")
    parser.add_argument("--sleep", type=float, default=0.2)
    parser.add_argument("--timeout", type=int, default=20)
    return parser.parse_args()


def main() -> None:
    configure_output_encoding()
    args = parse_args()
    report = backfill_coordinates(
        input_path=Path(args.input),
        output_path=Path(args.output),
        report_path=Path(args.report),
        cache_path=Path(args.cache),
        limit=args.limit,
        max_new_geocodes=args.max_new_geocodes,
        cache_only=args.cache_only,
        sleep_seconds=args.sleep,
        timeout=args.timeout,
    )
    print(json.dumps(report, ensure_ascii=False, indent=2), flush=True)


if __name__ == "__main__":
    main()
