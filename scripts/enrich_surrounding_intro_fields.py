# -*- coding: utf-8 -*-
"""
Append customer-readable surrounding/intro fields to merged auction tables.

The crawler keeps source text in JD-compatible columns. This post-processing
step preserves the original schema and appends derived columns that summarize
the long-text fields without removing traceability to the original text.
"""
from __future__ import annotations

import argparse
import html
import json
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Sequence, Tuple

import xlsxwriter
from openpyxl import load_workbook


INTRO_COL = "标的物介绍文本"
NOTICE_COL = "竞买公告文本"
ATTACH_INDEX_COL = "附件索引原文"

ADDED_COLUMNS = [
    "标的物介绍摘要_整理",
    "周边配套_整理",
    "交通情况_整理",
    "教育医疗商业_整理",
    "居住环境_整理",
    "整理命中关键词",
    "整理依据字段",
    "整理状态",
]


def configure_output_encoding() -> None:
    for stream_name in ("stdout", "stderr"):
        stream = getattr(sys, stream_name, None)
        if hasattr(stream, "reconfigure"):
            try:
                stream.reconfigure(encoding="utf-8", errors="backslashreplace")
            except Exception:
                pass


def print_report(report: Dict[str, Any]) -> None:
    text = json.dumps(report, ensure_ascii=False, indent=2)
    try:
        print(text, flush=True)
    except UnicodeEncodeError:
        print(text.encode("utf-8", errors="backslashreplace").decode("utf-8"), flush=True)

SECTION_HEADINGS = [
    "标的物估值",
    "标的评估总价",
    "评估总价",
    "费用总价",
    "税费情况",
    "其他费用情况",
    "其他介绍",
    "其他说明",
    "附件下载",
    "特别提醒",
    "竞买须知",
    "竞买公告",
    "拍卖公告",
    "权利限制情况",
    "提供的文件",
    "拟提供的文件",
    "拍品介绍",
    "标的物介绍",
    "房屋户型",
    "房屋楼层",
    "房屋朝向",
    "装修情况",
    "建筑总面积",
    "建筑面积",
    "土地总面积",
    "房产年龄",
    "成新率",
]

SURROUNDING_LABELS = [
    "周边配套",
    "周边情况",
    "周边环境",
    "周边状况",
    "区位状况",
    "区位情况",
    "配套情况",
    "生活配套",
    "公共配套",
]

TRAFFIC_WORDS = ["交通", "公交", "地铁", "车站", "高速", "道路", "路网", "出行", "便捷"]
EDU_MED_BIZ_WORDS = [
    "学校",
    "幼儿园",
    "小学",
    "中学",
    "医院",
    "卫生所",
    "诊所",
    "银行",
    "商场",
    "超市",
    "市场",
    "商业",
    "百货",
    "酒店",
]
ENV_WORDS = ["绿地", "环境", "景观", "居住氛围", "居住环境", "自然及人文环境", "成熟社区"]
STRONG_SURROUNDING_WORDS = sorted(
    set(SURROUNDING_LABELS + TRAFFIC_WORDS + EDU_MED_BIZ_WORDS + ENV_WORDS)
)
ALL_KEYWORDS = sorted(set(SURROUNDING_LABELS + TRAFFIC_WORDS + EDU_MED_BIZ_WORDS + ENV_WORDS))
WEAK_ONLY_WORDS = ["商业用房", "住宅用途", "住宅用房", "办公用房", "用途"]


def is_invalid_surrounding(value: str) -> bool:
    text = compact(value, 800)
    if not text:
        return True
    lower = text.lower()
    if any(ext in lower for ext in [".pdf", ".doc", ".docx", ".xls", ".xlsx", "attachmentaddress", "https://"]):
        return True
    stripped = re.sub(r"[。；;，,\s（）()详见]", "", text)
    if stripped in {"评估报告", "报告", "无", "不详"}:
        return True
    if text in {"见评估报告。", "详见评估报告。", "详见评估报告", "见评估报告"}:
        return True
    if any(text == word or text == f"{word}。" for word in WEAK_ONLY_WORDS):
        return True
    return False


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = html.unescape(str(value))
    text = text.replace("\r", "\n").replace("\u00a0", " ")
    text = re.sub(r"[ \t]{2,}", " ", text)
    text = re.sub(r"\n[ \t]+", "\n", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def compact(value: str, limit: int = 600) -> str:
    text = re.sub(r"\s+", " ", clean_text(value)).strip()
    if len(text) > limit:
        return text[: limit - 1] + "…"
    return text


def split_sentences(text: str) -> List[str]:
    normalized = clean_text(text)
    normalized = re.sub(r"\n+", "。", normalized)
    parts = re.split(r"(?<=[。！？；;])|(?<=\.)\s+", normalized)
    out: List[str] = []
    for part in parts:
        item = compact(part, 260)
        if item and item not in out:
            out.append(item)
    return out


def has_meaningful_surrounding(text: str) -> bool:
    source = clean_text(text)
    if not source:
        return False
    return any(word in source for word in STRONG_SURROUNDING_WORDS)


def extract_label_section(text: str, labels: Sequence[str], limit: int = 500) -> str:
    source = clean_text(text)
    if not source:
        return ""
    lines = [line.strip() for line in source.splitlines()]
    for idx, line in enumerate(lines):
        line_key = line.rstrip("：:")
        if line_key in labels or any(line_key.endswith(label) for label in labels):
            collected: List[str] = []
            for nxt in lines[idx + 1 :]:
                key = nxt.rstrip("：:")
                if not nxt:
                    if collected:
                        break
                    continue
                if key in SECTION_HEADINGS and collected:
                    break
                if key in labels and collected:
                    break
                collected.append(nxt)
                if len(" ".join(collected)) >= limit:
                    break
            return compact(" ".join(collected), limit)

    label_re = "|".join(re.escape(label) for label in labels)
    stop_re = "|".join(re.escape(h) for h in SECTION_HEADINGS if h not in labels)
    match = re.search(rf"(?:{label_re})[：:\s]+(.+?)(?=(?:{stop_re})[：:\s]|\n\n|$)", source, re.S)
    if match:
        return compact(match.group(1), limit)
    return ""


def select_keyword_sentences(text: str, words: Sequence[str], limit: int = 500, max_count: int = 4) -> str:
    selected: List[str] = []
    for sentence in split_sentences(text):
        if any(word in sentence for word in words):
            selected.append(sentence)
        if len(selected) >= max_count:
            break
    return compact(" ".join(selected), limit)


def intro_summary(text: str, limit: int = 700) -> str:
    source = clean_text(text)
    if not source:
        return ""
    section = extract_label_section(source, ["标的物介绍", "拍品介绍", "标的调查情况表"], limit=limit)
    if section:
        return section
    return compact(source, limit)


def derive_fields(row_map: Dict[str, Any]) -> Dict[str, str]:
    intro = clean_text(row_map.get(INTRO_COL))
    notice = clean_text(row_map.get(NOTICE_COL))
    attach = clean_text(row_map.get(ATTACH_INDEX_COL))
    combined = "\n".join(part for part in [intro, notice, attach] if part)
    prose = "\n".join(part for part in [intro, notice] if part)

    surrounding = extract_label_section(prose, SURROUNDING_LABELS, limit=650)
    if not surrounding:
        surrounding = select_keyword_sentences(prose, STRONG_SURROUNDING_WORDS, limit=650)
    if is_invalid_surrounding(surrounding):
        surrounding = ""

    traffic = select_keyword_sentences(surrounding or prose, TRAFFIC_WORDS, limit=450, max_count=3)
    edu_med_biz = select_keyword_sentences(surrounding or prose, EDU_MED_BIZ_WORDS, limit=450, max_count=4)
    environment = select_keyword_sentences(surrounding or prose, ENV_WORDS, limit=450, max_count=3)
    if is_invalid_surrounding(traffic):
        traffic = ""
    if is_invalid_surrounding(edu_med_biz):
        edu_med_biz = ""
    if is_invalid_surrounding(environment):
        environment = ""

    hit_keywords = [word for word in ALL_KEYWORDS if word in combined]
    basis = []
    if intro:
        basis.append(INTRO_COL)
    if notice:
        basis.append(NOTICE_COL)
    if attach:
        basis.append(ATTACH_INDEX_COL)

    has_any = any([surrounding, traffic, edu_med_biz, environment])
    status = "已整理" if has_any else ("无有效周边内容" if combined else "无可整理原文")
    return {
        "标的物介绍摘要_整理": intro_summary(intro or combined),
        "周边配套_整理": surrounding,
        "交通情况_整理": traffic,
        "教育医疗商业_整理": edu_med_biz,
        "居住环境_整理": environment,
        "整理命中关键词": "、".join(hit_keywords[:30]),
        "整理依据字段": "、".join(basis),
        "整理状态": status,
    }


def iter_rows(path: Path) -> Tuple[List[str], Iterable[Tuple[int, List[Any]]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [clean_text(value) for value in header_row]

    def generator() -> Iterable[Tuple[int, List[Any]]]:
        try:
            for row_index, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                yield row_index, list(values)
        finally:
            workbook.close()

    return headers, generator()


def enrich_excel(input_path: Path, output_path: Path, report_path: Path, tmpdir: Path | None = None) -> Dict[str, Any]:
    headers, rows = iter_rows(input_path)
    row_indexes = {name: idx for idx, name in enumerate(headers)}
    missing_required = [col for col in [INTRO_COL, NOTICE_COL, ATTACH_INDEX_COL] if col not in row_indexes]
    if missing_required:
        raise ValueError(f"missing required columns: {missing_required}")

    new_columns = [col for col in ADDED_COLUMNS if col not in headers]
    output_headers = list(headers) + new_columns
    output_path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = output_path.with_suffix(output_path.suffix + ".tmp.xlsx")
    workbook_options = {"strings_to_urls": False, "constant_memory": True, "use_zip64": True}
    if tmpdir is not None:
        tmpdir.mkdir(parents=True, exist_ok=True)
        workbook_options["tmpdir"] = str(tmpdir)
    workbook = xlsxwriter.Workbook(str(tmp_path), workbook_options)
    worksheet = workbook.add_worksheet("Sheet1")
    for col, header in enumerate(output_headers):
        worksheet.write(0, col, header)

    total = 0
    status_counts: Dict[str, int] = {}
    filled_counts = {col: 0 for col in ADDED_COLUMNS}
    samples: List[Dict[str, Any]] = []
    for out_row, (_source_row_index, values) in enumerate(rows, start=1):
        total += 1
        row_map = {name: values[idx] if idx < len(values) else None for name, idx in row_indexes.items()}
        derived = derive_fields(row_map)
        status = derived["整理状态"]
        status_counts[status] = status_counts.get(status, 0) + 1
        for col in ADDED_COLUMNS:
            if derived.get(col):
                filled_counts[col] += 1
        if len(samples) < 10 and status == "已整理":
            samples.append(
                {
                    "row": total + 1,
                    "id": clean_text(row_map.get("标的物ID")),
                    "source": clean_text(row_map.get("publishSource")),
                    "周边配套_整理": derived["周边配套_整理"],
                }
            )

        output_values = list(values) + [derived[col] for col in new_columns]
        for col, value in enumerate(output_values):
            worksheet.write(out_row, col, "" if value is None else value)

    workbook.close()
    tmp_path.replace(output_path)

    report = {
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "input": str(input_path),
        "output": str(output_path),
        "original_columns": len(headers),
        "output_columns": len(output_headers),
        "rows": total,
        "added_columns": new_columns,
        "status_counts": status_counts,
        "filled_counts": filled_counts,
        "samples": samples,
    }
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    return report


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Append organized surrounding and intro fields.")
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--report", required=True)
    parser.add_argument("--tmpdir")
    return parser.parse_args()


def main() -> None:
    configure_output_encoding()
    args = parse_args()
    report = enrich_excel(
        Path(args.input),
        Path(args.output),
        Path(args.report),
        Path(args.tmpdir) if args.tmpdir else None,
    )
    print_report(report)


if __name__ == "__main__":
    main()
