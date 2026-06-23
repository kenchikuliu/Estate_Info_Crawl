# -*- coding: utf-8 -*-
"""
Backfill Alibaba judicial-auction detail fields through public detail-ext APIs.

The normal PC detail page often redirects to Taobao login, but several lazy
detail endpoints are still public. This script uses those endpoints directly,
writes one JSON line per finished item, and exports a merged Excel file.
"""
from __future__ import annotations

import argparse
import json
import math
import re
import sys
import threading
import time
from concurrent.futures import FIRST_COMPLETED, ThreadPoolExecutor, as_completed, wait
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Set, Tuple
from urllib.parse import urljoin

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from utils.jd_detail_parser import (  # noqa: E402
    clean_text,
    extract_labeled_fields,
    extract_region_fields,
    extract_rights_status_text,
    parse_intro_sections,
    parse_survey_table_rows,
    postprocess_structured_fields,
)


DEFAULT_INPUT = r"output\ali_sf_gd_full_h5_20260611\阿里法拍房_广东_全量索引.xlsx"
DEFAULT_OUTPUT = r"output\ali_sf_gd_full_h5_20260611\阿里法拍房_广东_详情回填_API.xlsx"
DEFAULT_JSONL = r"output\ali_sf_gd_full_h5_20260611\阿里法拍房_广东_详情回填_API.jsonl"
DEFAULT_CHECKPOINT = r"output\ali_sf_gd_full_h5_20260611\阿里法拍房_广东_详情回填_API.checkpoint.json"
ILLEGAL_XML_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F\x7F-\x9F]")
JSONP_PAYLOAD_RE = re.compile(r"^[^(]*\((.*)\)\s*;?\s*$", re.S)
FETCH_COLUMN_CANDIDATES = [
    "标的物ID",
    "链接",
    "详情页URL",
    "itemId",
    "item_id",
    "id",
    "类目ID",
    "categoryId",
    "cat_id",
    "标题",
    "标的名称",
    "标的物名称",
    "完整地址",
    "格式化地址",
    "省",
    "省份",
    "市",
    "城市",
    "区县",
    "经度",
    "纬度",
    "longitude",
    "latitude",
    "lng",
    "lat",
    "法院",
    "courtName",
    "处置机构",
    "处置法院",
    "竞价状态",
    "成交状态",
    "status",
    "当前价_元",
    "当前价",
    "起拍价_元",
    "起拍价",
    "保证金_元",
    "保证金",
    "评估价_元",
    "评估价",
    "市场价_元",
    "市场价",
    "围观次数",
    "提醒人数",
    "出价次数",
    "报名人数",
    "延时次数",
    "成交价/获拍价_元",
    "成交时间",
    "是否成交",
    "是否流拍",
    "标的所有人",
    "被执行人",
]

DETAIL_COLUMNS = [
    "详情抓取时间",
    "详情抓取状态",
    "详情抓取错误",
    "详情页URL",
    "详情标题",
    "标的名称",
    "标的物名称",
    "权证情况",
    "标的所有人",
    "被执行人/标的所有人",
    "详情地址",
    "详情省",
    "详情市",
    "详情区县",
    "经度",
    "纬度",
    "经纬度来源",
    "法院/处置机构",
    "拍卖状态_详情",
    "auctionType_详情",
    "是否成交",
    "是否流拍",
    "成交价/获拍价_元",
    "成交时间",
    "当前价_详情_元",
    "起拍价_详情_元",
    "保证金_详情_元",
    "评估价_详情_元",
    "市场价_详情_元",
    "出价次数_详情",
    "围观次数_详情",
    "报名人数_详情",
    "延时次数_详情",
    "建筑面积",
    "房屋用途",
    "房屋类型",
    "所在层",
    "总层数",
    "竣工时间",
    "购买时间",
    "土地性质",
    "土地用途",
    "使用期限",
    "权利来源",
    "所有权来源",
    "钥匙/占用情况",
    "腾空情况",
    "户籍/工商注册",
    "欠费情况",
    "提供文件",
    "权利限制状况及抵押状况",
    "房屋权属状况",
    "土地权属状况",
    "附件数量",
    "附件名称",
    "附件链接",
    "附件ID",
    "附件索引原文",
    "标的物介绍文本",
]

thread_local = threading.local()


def make_session() -> requests.Session:
    session = requests.Session()
    session.trust_env = False
    session.headers.update(
        {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/148.0.0.0 Safari/537.36"
            ),
            "Accept": "application/json,text/javascript,text/html,*/*",
            "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
            "Connection": "close",
        }
    )
    return session


def get_session() -> requests.Session:
    session = getattr(thread_local, "session", None)
    if session is None:
        session = make_session()
        thread_local.session = session
    return session


def sanitize_excel_value(value: Any) -> Any:
    if isinstance(value, str):
        return ILLEGAL_XML_RE.sub("", value)
    return value


def sanitize_dataframe_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    return df.apply(lambda col: col.map(sanitize_excel_value))


def dump_json(value: Any) -> str:
    if value in ("", None, [], {}):
        return ""
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"))


def normalize_id(value: Any) -> str:
    if value in ("", None):
        return ""
    text = str(value).strip()
    if not text or text.lower() == "nan":
        return ""
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    match = re.search(r"sf_item/(\d+)\.htm", text)
    if match:
        return match.group(1)
    match = re.search(r"[?&](?:id|item_id|itemId)=(\d+)", text)
    if match:
        return match.group(1)
    return text if text.isdigit() else ""


def normalize_url(value: Any, base_url: str = "https://sf-item.taobao.com/") -> str:
    text = clean_text(str(value or ""))
    if not text:
        return ""
    if text.startswith("//"):
        return "https:" + text
    if text.startswith("/"):
        return urljoin(base_url, text)
    return text


def first_non_empty(*values: Any) -> Any:
    for value in values:
        if value in (None, ""):
            continue
        if isinstance(value, float) and pd.isna(value):
            continue
        return value
    return ""


def join_unique(values: Iterable[Any]) -> str:
    result: List[str] = []
    seen = set()
    for value in values:
        text = clean_text(str(value or ""))
        if not text or text in seen:
            continue
        seen.add(text)
        result.append(text)
    return "；".join(result)


def to_int(value: Any, default: int = 0) -> int:
    try:
        if value in ("", None):
            return default
        return int(float(value))
    except Exception:
        return default


def to_float_or_blank(value: Any) -> Any:
    try:
        if value in ("", None):
            return ""
        number = float(str(value).replace(",", ""))
    except Exception:
        return ""
    if math.isnan(number):
        return ""
    return number


def taobao_bid_price_yuan(value: Any, reference_yuan: Any = "") -> Any:
    number = to_float_or_blank(value)
    if number == "":
        return ""
    reference = to_float_or_blank(reference_yuan)
    if reference != "" and reference > 0:
        return number / 100 if number > reference * 10 else number
    return number / 100


def ts_to_text(value: Any) -> str:
    if value in ("", None):
        return ""
    try:
        number = float(value)
    except (TypeError, ValueError):
        return ""
    if number <= 0:
        return ""
    if number > 10_000_000_000:
        number = number / 1000
    return datetime.fromtimestamp(number).strftime("%Y-%m-%d %H:%M:%S")


def truncate_text(text: str, limit: int) -> str:
    normalized = clean_text(text)
    if limit <= 0 or len(normalized) <= limit:
        return normalized
    return normalized[:limit] + "..."


def parse_jsonish(text: str) -> Any:
    raw = text.strip()
    match = JSONP_PAYLOAD_RE.match(raw)
    if match:
        raw = match.group(1)
    try:
        return json.loads(raw)
    except Exception:
        return None


def response_text(response: requests.Response) -> str:
    if not response.encoding or response.encoding.lower() == "iso-8859-1":
        response.encoding = response.apparent_encoding or "utf-8"
    return response.text


def api_get(url: str, item_id: str, retries: int, timeout: int, sleep_seconds: float = 0.8) -> Any:
    last_error: Optional[Exception] = None
    headers = {"Referer": f"https://sf-item.taobao.com/sf_item/{item_id}.htm"}
    for attempt in range(1, retries + 1):
        try:
            session = get_session()
            response = session.get(url, headers=headers, timeout=timeout)
            response.raise_for_status()
            text = response_text(response)
            payload = parse_jsonish(text)
            if payload is None:
                return text
            return payload
        except Exception as exc:  # noqa: BLE001
            last_error = exc
            time.sleep(sleep_seconds * attempt)
    raise RuntimeError(str(last_error))


def html_to_text_and_rows(html: str) -> Tuple[str, List[List[str]]]:
    if not html:
        return "", []
    soup = BeautifulSoup(html, "lxml")
    for tag in soup(["script", "style"]):
        tag.decompose()
    rows: List[List[str]] = []
    for tr in soup.find_all("tr"):
        cells = [clean_text(cell.get_text(" ", strip=True)) for cell in tr.find_all(["td", "th"])]
        cells = [cell for cell in cells if cell]
        if cells:
            rows.append(cells)
    text = clean_text(soup.get_text("\n"))
    return text, rows


def parse_description_fields(html: str) -> Tuple[Dict[str, str], str]:
    text, rows = html_to_text_and_rows(html)
    fields: Dict[str, str] = {}

    sections = parse_intro_sections(text)
    if sections.get("拍品名称"):
        fields["标的物名称"] = sections["拍品名称"]
    if sections.get("权证情况"):
        fields["权证情况"] = sections["权证情况"]
    if sections.get("拍品所有人"):
        fields["被执行人"] = sections["拍品所有人"]
    if sections.get("成交后提供的文件"):
        fields["提供文件"] = sections["成交后提供的文件"]

    fields.update(extract_labeled_fields(text))
    fields.update(parse_survey_table_rows(rows))
    fields.update(extract_rights_status_text(text))
    return postprocess_structured_fields(fields), text


def extract_content(payload: Any) -> str:
    if isinstance(payload, str):
        return payload
    if not isinstance(payload, dict):
        return ""
    data = payload.get("data")
    if isinstance(data, dict):
        for key in ["content", "desc", "html", "value"]:
            if isinstance(data.get(key), str):
                return data[key]
    if isinstance(data, str):
        return data
    for key in ["content", "desc", "html", "value"]:
        if isinstance(payload.get(key), str):
            return payload[key]
    return ""


def normalize_attachment(item: Dict[str, Any]) -> Dict[str, str]:
    title = first_non_empty(item.get("title"), item.get("name"), item.get("fileName"), item.get("attachmentName"))
    attach_id = first_non_empty(item.get("id"), item.get("fileId"), item.get("attachmentId"))
    url = first_non_empty(item.get("url"), item.get("href"), item.get("downloadUrl"), item.get("attachmentAddress"))
    return {
        "title": clean_text(str(title or "")),
        "id": clean_text(str(attach_id or "")),
        "fileType": clean_text(str(item.get("fileType", ""))),
        "url": normalize_url(url),
    }


def extract_attachments(payload: Any) -> List[Dict[str, str]]:
    if not isinstance(payload, dict):
        return []
    data = payload.get("data")
    raw: Any = []
    if isinstance(data, dict):
        raw = data.get("attaches") or data.get("attachments") or data.get("files") or []
    elif isinstance(data, list):
        raw = data
    if not isinstance(raw, list):
        return []
    return [normalize_attachment(item) for item in raw if isinstance(item, dict)]


def best_bid_record(bid_records: Any) -> Dict[str, Any]:
    if not isinstance(bid_records, list):
        return {}
    records = [item for item in bid_records if isinstance(item, dict)]
    if not records:
        return {}
    deal_records = [item for item in records if clean_text(str(item.get("bidStatus") or "")) == "deal"]
    candidates = deal_records or records
    return max(candidates, key=lambda item: (to_float_or_blank(item.get("bidPrice")) or 0, to_int(item.get("bidTime"), 0)))


def summarize_bid_detail(payload: Any, base_row: Dict[str, Any]) -> Dict[str, Any]:
    data = payload.get("data") if isinstance(payload, dict) else {}
    data = data if isinstance(data, dict) else {}
    bid_count = first_non_empty(data.get("bidCount"), base_row.get("出价次数"))
    apply_count = first_non_empty(data.get("applyCnt"), base_row.get("报名人数"))
    bid = best_bid_record(data.get("bidRecords"))
    base_deal_price = first_non_empty(base_row.get("成交价/获拍价_元"), base_row.get("当前价_元"), base_row.get("当前价"))
    deal_price = first_non_empty(taobao_bid_price_yuan(bid.get("bidPrice"), base_deal_price), base_row.get("成交价/获拍价_元"))
    deal_time = first_non_empty(ts_to_text(bid.get("bidTime")), base_row.get("成交时间"))
    sold = bool(bid and clean_text(str(bid.get("bidStatus") or "")) == "deal")
    base_sold = clean_text(str(base_row.get("是否成交") or ""))
    base_unsold = clean_text(str(base_row.get("是否流拍") or ""))
    return {
        "auctionType_详情": data.get("auctionType", ""),
        "是否成交": "是" if sold else (base_sold or "否"),
        "是否流拍": base_unsold or ("否" if sold else ""),
        "成交价/获拍价_元": deal_price if sold or base_sold == "是" else "",
        "成交时间": deal_time if sold or base_sold == "是" else "",
        "出价次数_详情": bid_count,
        "报名人数_详情": apply_count,
        "延时次数_详情": first_non_empty(data.get("delayCnt"), base_row.get("延时次数")),
    }


def extract_region_from_address(address: str) -> Dict[str, str]:
    try:
        return extract_region_fields(address)
    except Exception:
        return {"省": "", "市": "", "区县": ""}


def address_fields(base_row: Dict[str, Any], fields: Dict[str, str]) -> Dict[str, str]:
    address = first_non_empty(
        fields.get("标的物名称"),
        base_row.get("完整地址"),
        base_row.get("格式化地址"),
        base_row.get("标题"),
    )
    region = extract_region_from_address(str(address or ""))
    return {
        "详情地址": clean_text(str(address or "")),
        "详情省": first_non_empty(base_row.get("省"), base_row.get("省份"), region.get("省")),
        "详情市": first_non_empty(base_row.get("市"), base_row.get("城市"), region.get("市")),
        "详情区县": first_non_empty(base_row.get("区县"), region.get("区县")),
    }


def coordinate_fields(base_row: Dict[str, Any]) -> Dict[str, Any]:
    lng = first_non_empty(base_row.get("经度"), base_row.get("longitude"), base_row.get("lng"))
    lat = first_non_empty(base_row.get("纬度"), base_row.get("latitude"), base_row.get("lat"))
    return {
        "经度": lng,
        "纬度": lat,
        "经纬度来源": "阿里索引API" if lng and lat else "",
    }


def fetch_detail(
    base_row: Dict[str, Any],
    include_description: bool,
    include_attachments: bool,
    include_bid: bool,
    raw_text_limit: int,
    retries: int,
    timeout: int,
) -> Dict[str, Any]:
    item_id = normalize_id(base_row.get("标的物ID") or base_row.get("链接"))
    errors: List[str] = []
    desc_html = ""
    desc_fields: Dict[str, str] = {}
    desc_text = ""
    attachments: List[Dict[str, str]] = []
    bid_summary: Dict[str, Any] = {}

    if include_description:
        desc_url = f"https://detail-ext.taobao.com/json/get_auction_desc_content.do?item_id={item_id}&_input_charset=UTF-8"
        try:
            payload = api_get(desc_url, item_id, retries=retries, timeout=timeout)
            desc_html = extract_content(payload)
            if desc_html:
                desc_fields, desc_text = parse_description_fields(desc_html)
            else:
                errors.append("description:empty")
        except Exception as exc:  # noqa: BLE001
            errors.append(f"description:{exc}")

    if include_attachments:
        cat_id = clean_text(str(first_non_empty(base_row.get("类目ID"), base_row.get("categoryId"), base_row.get("cat_id")) or ""))
        attach_url = (
            f"https://detail-ext.taobao.com/json/get_item_attach.do?item_id={item_id}&id={item_id}"
            f"&cat_id={cat_id}&preview_key=&_input_charset=UTF-8"
        )
        try:
            payload = api_get(attach_url, item_id, retries=retries, timeout=timeout)
            attachments = extract_attachments(payload)
        except Exception as exc:  # noqa: BLE001
            errors.append(f"attachments:{exc}")

    if include_bid:
        bid_url = f"https://sf-item.taobao.com/json/get_bid_detail.htm?id={item_id}"
        try:
            payload = api_get(bid_url, item_id, retries=retries, timeout=timeout)
            bid_summary = summarize_bid_detail(payload, base_row)
        except Exception as exc:  # noqa: BLE001
            errors.append(f"bid:{exc}")

    owner = first_non_empty(desc_fields.get("被执行人"), base_row.get("标的所有人"), base_row.get("被执行人"))
    title = first_non_empty(desc_fields.get("标的物名称"), base_row.get("标题"), base_row.get("标的名称"))
    row: Dict[str, Any] = {
        "标的物ID": item_id,
        "详情抓取时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "详情抓取状态": "成功" if not errors else ("部分成功" if desc_text or attachments or bid_summary else "失败"),
        "详情抓取错误": " | ".join(errors),
        "详情页URL": f"https://sf-item.taobao.com/sf_item/{item_id}.htm",
        "详情标题": clean_text(str(title or "")),
        "标的名称": clean_text(str(title or "")),
        "标的物名称": clean_text(str(first_non_empty(desc_fields.get("标的物名称"), title) or "")),
        "权证情况": desc_fields.get("权证情况", ""),
        "标的所有人": owner,
        "被执行人/标的所有人": owner,
        **address_fields(base_row, desc_fields),
        **coordinate_fields(base_row),
        "法院/处置机构": first_non_empty(base_row.get("法院"), base_row.get("courtName"), base_row.get("处置机构"), base_row.get("处置法院")),
        "拍卖状态_详情": first_non_empty(base_row.get("竞价状态"), base_row.get("成交状态"), base_row.get("status")),
        "当前价_详情_元": first_non_empty(base_row.get("当前价_元"), base_row.get("当前价")),
        "起拍价_详情_元": first_non_empty(base_row.get("起拍价_元"), base_row.get("起拍价")),
        "保证金_详情_元": first_non_empty(base_row.get("保证金_元"), base_row.get("保证金")),
        "评估价_详情_元": first_non_empty(base_row.get("评估价_元"), base_row.get("评估价")),
        "市场价_详情_元": first_non_empty(base_row.get("市场价_元"), base_row.get("市场价")),
        "围观次数_详情": first_non_empty(base_row.get("围观次数"), base_row.get("提醒人数")),
        "建筑面积": desc_fields.get("建筑面积", ""),
        "房屋用途": desc_fields.get("房屋用途", ""),
        "房屋类型": desc_fields.get("房屋类型", ""),
        "所在层": desc_fields.get("所在层", ""),
        "总层数": desc_fields.get("总层数", ""),
        "竣工时间": desc_fields.get("竣工时间", ""),
        "购买时间": desc_fields.get("购买时间", ""),
        "土地性质": desc_fields.get("土地性质", ""),
        "土地用途": desc_fields.get("土地用途", ""),
        "使用期限": desc_fields.get("使用期限", ""),
        "权利来源": desc_fields.get("权利来源", ""),
        "所有权来源": desc_fields.get("所有权来源", ""),
        "钥匙/占用情况": desc_fields.get("钥匙", ""),
        "腾空情况": desc_fields.get("腾空情况", ""),
        "户籍/工商注册": desc_fields.get("户籍注册", ""),
        "欠费情况": desc_fields.get("欠费情况", ""),
        "提供文件": desc_fields.get("提供文件", ""),
        "权利限制状况及抵押状况": desc_fields.get("权利限制状况及抵押状况", ""),
        "房屋权属状况": desc_fields.get("房屋权属状况", ""),
        "土地权属状况": desc_fields.get("土地权属状况", ""),
        "附件数量": len(attachments),
        "附件名称": join_unique(item.get("title") for item in attachments),
        "附件链接": join_unique(item.get("url") for item in attachments),
        "附件ID": join_unique(item.get("id") for item in attachments),
        "附件索引原文": dump_json(attachments),
        "标的物介绍文本": truncate_text(desc_text, raw_text_limit),
    }
    row.update(bid_summary)
    return row


def find_column(columns: Sequence[str], candidates: Sequence[str]) -> str:
    for candidate in candidates:
        if candidate in columns:
            return candidate
    return ""


def read_index(path: Path, start: int = 0, limit: int = 0, usecols: Optional[Sequence[str]] = None) -> pd.DataFrame:
    read_kwargs: Dict[str, Any] = {}
    if start > 0:
        read_kwargs["skiprows"] = range(1, start + 1)
    if limit and limit > 0:
        read_kwargs["nrows"] = limit
    if usecols is not None:
        wanted = set(usecols)
        read_kwargs["usecols"] = lambda col: col in wanted
    df = pd.read_excel(path, **read_kwargs).fillna("").astype(object)
    id_col = find_column(list(df.columns), ["标的物ID", "拍品ID", "itemId", "item_id", "id"])
    link_col = find_column(list(df.columns), ["链接", "详情页URL", "详情链接", "url", "URL"])
    if id_col:
        df["标的物ID"] = df[id_col].map(normalize_id)
    elif link_col:
        df["标的物ID"] = df[link_col].map(normalize_id)
    else:
        raise ValueError("Input Excel must contain 标的物ID or 链接.")
    if "链接" not in df.columns:
        df["链接"] = df["标的物ID"].map(lambda value: f"https://sf-item.taobao.com/sf_item/{value}.htm" if value else "")
    return df[df["标的物ID"].astype(str).str.len() > 0].copy()


def ordered_unique_rows(df: pd.DataFrame) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    seen: Set[str] = set()
    for row in df.to_dict("records"):
        item_id = normalize_id(row.get("标的物ID") or row.get("链接"))
        if not item_id or item_id in seen:
            continue
        row = dict(row)
        row["标的物ID"] = item_id
        rows.append(row)
        seen.add(item_id)
    return rows


def iter_index_rows_stream(
    path: Path,
    start: int = 0,
    limit: int = 0,
    usecols: Optional[Sequence[str]] = None,
) -> Iterable[Dict[str, Any]]:
    wanted = set(usecols or [])
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        headers = next(rows_iter, None)
        if not headers:
            return
        header_names = [clean_text(str(value or "")) for value in headers]
        selected_indexes = [
            index
            for index, name in enumerate(header_names)
            if name and (not wanted or name in wanted)
        ]
        produced = 0
        for row_number, values in enumerate(rows_iter):
            data_index = row_number
            if data_index < start:
                continue
            row: Dict[str, Any] = {}
            for index in selected_indexes:
                if index >= len(values):
                    continue
                value = values[index]
                if value in (None, ""):
                    continue
                row[header_names[index]] = value
            item_id = normalize_id(
                first_non_empty(
                    row.get("标的物ID"),
                    row.get("拍品ID"),
                    row.get("itemId"),
                    row.get("item_id"),
                    row.get("id"),
                    row.get("链接"),
                    row.get("详情页URL"),
                )
            )
            if not item_id:
                continue
            row["标的物ID"] = item_id
            if "链接" not in row:
                row["链接"] = f"https://sf-item.taobao.com/sf_item/{item_id}.htm"
            yield row
            produced += 1
            if limit and limit > 0 and produced >= limit:
                break
    finally:
        workbook.close()


def selected_rows_stream(
    input_path: Path,
    start: int,
    limit: int,
    usecols: Sequence[str],
    wanted_ids: Optional[Set[str]],
    completed_before: Set[str],
) -> Iterable[Dict[str, Any]]:
    seen: Set[str] = set()
    for row in iter_index_rows_stream(input_path, start=start, limit=limit, usecols=usecols):
        item_id = normalize_id(row.get("标的物ID"))
        if not item_id or item_id in seen:
            continue
        seen.add(item_id)
        if wanted_ids is not None and item_id not in wanted_ids:
            continue
        if item_id in completed_before:
            continue
        yield row


def read_jsonl(path: Path) -> List[Dict[str, Any]]:
    if not path.exists():
        return []
    rows: List[Dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if not line:
                continue
            try:
                item = json.loads(line)
            except Exception:
                continue
            if isinstance(item, dict):
                rows.append(item)
    return rows


def load_completed_ids(path: Path) -> Set[str]:
    return {
        normalize_id(row.get("标的物ID"))
        for row in read_jsonl(path)
        if normalize_id(row.get("标的物ID")) and row.get("详情抓取状态") in {"成功", "部分成功"}
    }


def append_jsonl(path: Path, row: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(row, ensure_ascii=False, separators=(",", ":")) + "\n")


def write_checkpoint(path: Path, payload: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = dict(payload)
    payload["updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def export_excel(index_df: pd.DataFrame, jsonl_path: Path, output_path: Path, completed_only: bool) -> int:
    detail_rows = read_jsonl(jsonl_path)
    if not detail_rows:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        pd.DataFrame().to_excel(output_path, index=False)
        return 0

    detail_df = pd.DataFrame(detail_rows).fillna("").astype(object)
    detail_df["标的物ID"] = detail_df["标的物ID"].map(normalize_id)
    detail_df["_detail_order"] = range(len(detail_df))
    detail_df = detail_df.sort_values("_detail_order").drop_duplicates("标的物ID", keep="last").drop(columns=["_detail_order"])

    base = index_df.copy()
    base["标的物ID"] = base["标的物ID"].map(normalize_id)
    duplicate_detail_cols = [col for col in detail_df.columns if col in base.columns and col != "标的物ID"]
    detail_df = detail_df.drop(columns=duplicate_detail_cols)
    merged = base.merge(detail_df, on="标的物ID", how="inner" if completed_only else "left")

    ordered_cols = list(base.columns)
    ordered_cols.extend(col for col in DETAIL_COLUMNS if col in merged.columns and col not in ordered_cols)
    ordered_cols.extend(col for col in merged.columns if col not in ordered_cols)
    merged = sanitize_dataframe_for_excel(merged[ordered_cols])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = output_path.with_suffix(output_path.suffix + ".tmp.xlsx")
    with pd.ExcelWriter(
        tmp_path,
        engine="xlsxwriter",
        engine_kwargs={"options": {"strings_to_urls": False}},
    ) as writer:
        merged.to_excel(writer, index=False)
    tmp_path.replace(output_path)
    return len(merged)


def detail_rows_by_id(jsonl_path: Path) -> Dict[str, Dict[str, Any]]:
    result: Dict[str, Dict[str, Any]] = {}
    for row in read_jsonl(jsonl_path):
        item_id = normalize_id(row.get("标的物ID"))
        if item_id:
            copied = dict(row)
            copied["标的物ID"] = item_id
            result[item_id] = copied
    return result


def export_excel_streaming(input_path: Path, jsonl_path: Path, output_path: Path, completed_only: bool) -> int:
    details = detail_rows_by_id(jsonl_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = output_path.with_suffix(output_path.suffix + ".tmp.xlsx")

    workbook = load_workbook(input_path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        headers_raw = next(rows_iter, None)
        if not headers_raw:
            pd.DataFrame().to_excel(tmp_path, index=False)
            tmp_path.replace(output_path)
            return 0

        base_headers = [clean_text(str(value or "")) for value in headers_raw]
        base_headers = [name or f"列{index + 1}" for index, name in enumerate(base_headers)]
        id_col = find_column(base_headers, ["标的物ID", "拍品ID", "itemId", "item_id", "id"])
        link_col = find_column(base_headers, ["链接", "详情页URL", "详情链接", "url", "URL"])
        detail_extra_cols: List[str] = []
        sample_detail_cols: Set[str] = set()
        for detail in details.values():
            sample_detail_cols.update(detail.keys())
        for col in DETAIL_COLUMNS:
            if col in sample_detail_cols and col not in base_headers:
                detail_extra_cols.append(col)
        for col in sorted(sample_detail_cols):
            if col not in {"标的物ID"} and col not in base_headers and col not in detail_extra_cols:
                detail_extra_cols.append(col)
        output_headers = base_headers + detail_extra_cols

        row_count = 0
        with pd.ExcelWriter(
            tmp_path,
            engine="xlsxwriter",
            engine_kwargs={"options": {"strings_to_urls": False}},
        ) as writer:
            worksheet = writer.book.add_worksheet("Sheet1")
            writer.sheets["Sheet1"] = worksheet
            for col_index, header in enumerate(output_headers):
                worksheet.write(0, col_index, sanitize_excel_value(header))

            output_row_index = 1
            for values in rows_iter:
                base_row = {
                    base_headers[index]: values[index] if index < len(values) and values[index] is not None else ""
                    for index in range(len(base_headers))
                }
                item_id = normalize_id(
                    first_non_empty(
                        base_row.get(id_col) if id_col else "",
                        base_row.get(link_col) if link_col else "",
                    )
                )
                detail = details.get(item_id, {})
                if completed_only and not detail:
                    continue
                row_values = [base_row.get(header, "") for header in base_headers]
                row_values.extend(detail.get(header, "") for header in detail_extra_cols)
                for col_index, value in enumerate(row_values):
                    worksheet.write(output_row_index, col_index, sanitize_excel_value(value))
                output_row_index += 1
                row_count += 1
    finally:
        workbook.close()

    tmp_path.replace(output_path)
    return row_count


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Backfill Ali Guangdong auction details through public APIs.")
    parser.add_argument("--input", default=DEFAULT_INPUT)
    parser.add_argument("--output", default=DEFAULT_OUTPUT)
    parser.add_argument("--jsonl-output", default=DEFAULT_JSONL)
    parser.add_argument(
        "--completed-jsonl",
        action="append",
        default=[],
        help="Additional JSONL file(s) whose successful IDs should be skipped while writing to --jsonl-output.",
    )
    parser.add_argument("--checkpoint", default=DEFAULT_CHECKPOINT)
    parser.add_argument("--ids", default="", help="Comma-separated item IDs or Ali detail URLs.")
    parser.add_argument("--start", type=int, default=0)
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--workers", type=int, default=6)
    parser.add_argument("--save-every", type=int, default=200)
    parser.add_argument("--export-every", type=int, default=0, help="0 means only export at the end.")
    parser.add_argument("--retries", type=int, default=4)
    parser.add_argument("--timeout", type=int, default=20)
    parser.add_argument("--raw-text-limit", type=int, default=2000)
    parser.add_argument("--full-read-for-fetch", action="store_true", help="Read all Excel columns before fetching. Slower on large files.")
    parser.add_argument("--no-resume", action="store_true", help="Do not skip IDs already present in JSONL.")
    parser.add_argument("--no-description", action="store_true")
    parser.add_argument("--no-attachments", action="store_true")
    parser.add_argument("--no-bid", action="store_true")
    parser.add_argument("--include-unfinished-index", action="store_true", help="Excel output keeps all index rows.")
    parser.add_argument("--no-export", action="store_true", help="Fetch JSONL/checkpoint only. Rebuild Excel later with --export-only.")
    parser.add_argument("--export-only", action="store_true", help="Only rebuild Excel from existing JSONL.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_path = Path(args.input)
    output_path = Path(args.output)
    jsonl_path = Path(args.jsonl_output)
    checkpoint_path = Path(args.checkpoint)

    if args.export_only:
        rows = export_excel_streaming(input_path, jsonl_path, output_path, completed_only=not args.include_unfinished_index)
        print(f"export_done output={output_path} rows={rows}", flush=True)
        return

    completed_before = load_completed_ids(jsonl_path) if not args.no_resume else set()
    if not args.no_resume:
        for completed_jsonl in args.completed_jsonl:
            completed_before.update(load_completed_ids(Path(completed_jsonl)))
    wanted_ids_list = [normalize_id(part.strip()) for part in args.ids.split(",") if normalize_id(part.strip())]
    wanted_ids = set(wanted_ids_list) if wanted_ids_list else None

    started_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    write_checkpoint(
        checkpoint_path,
        {
            "started_at": started_at,
            "input": str(input_path),
            "jsonl_output": str(jsonl_path),
            "excel_output": str(output_path),
            "selected": "stream" if not args.full_read_for_fetch else 0,
            "completed_existing": len(completed_before),
            "completed_this_run": 0,
            "failed_this_run": 0,
            "submitted": 0,
        },
    )

    completed_this_run = 0
    failed_this_run = 0
    submitted = 0
    submitted_ids: Set[str] = set()
    started = time.time()

    def handle_future(future: Any, base_row: Dict[str, Any]) -> None:
        nonlocal completed_this_run, failed_this_run
        item_id = normalize_id(base_row.get("标的物ID"))
        try:
            detail_row = future.result()
            if detail_row.get("详情抓取状态") == "失败":
                failed_this_run += 1
        except Exception as exc:  # noqa: BLE001
            failed_this_run += 1
            detail_row = {
                "标的物ID": item_id,
                "详情抓取时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "详情抓取状态": "失败",
                "详情抓取错误": str(exc),
                "详情页URL": f"https://sf-item.taobao.com/sf_item/{item_id}.htm",
            }
        append_jsonl(jsonl_path, detail_row)
        completed_this_run += 1

    def maybe_report(force: bool = False) -> None:
        if not force and completed_this_run % max(1, args.save_every) != 0:
            return
        elapsed = time.time() - started
        rate = completed_this_run / elapsed if elapsed > 0 else 0
        in_flight = submitted - completed_this_run
        write_checkpoint(
            checkpoint_path,
            {
                "started_at": started_at,
                "input": str(input_path),
                "jsonl_output": str(jsonl_path),
                "excel_output": str(output_path),
                "selected": "stream" if not args.full_read_for_fetch else submitted,
                "completed_existing": len(completed_before),
                "submitted": submitted,
                "completed_this_run": completed_this_run,
                "failed_this_run": failed_this_run,
                "in_flight": in_flight,
                "elapsed_seconds": round(elapsed, 1),
                "items_per_second": round(rate, 3),
            },
        )
        print(
            f"ali_api_detail_progress submitted={submitted} done={completed_this_run} "
            f"failed={failed_this_run} in_flight={in_flight} rate={rate:.2f}/s",
            flush=True,
        )

    fetch_usecols = None if args.full_read_for_fetch else FETCH_COLUMN_CANDIDATES
    if args.full_read_for_fetch:
        index_df = read_index(input_path, start=args.start, limit=args.limit, usecols=fetch_usecols)
        all_rows = ordered_unique_rows(index_df)
        if wanted_ids is not None:
            all_rows = [row for row in all_rows if row["标的物ID"] in wanted_ids]
        todo_iterable = (row for row in all_rows if row["标的物ID"] not in completed_before)
        print(
            f"ali_api_detail_plan input_rows={len(index_df)} unique_ids={len(all_rows)} "
            f"completed_existing={len(completed_before)} workers={args.workers}",
            flush=True,
        )
    else:
        todo_iterable = selected_rows_stream(
            input_path,
            start=args.start,
            limit=args.limit,
            usecols=FETCH_COLUMN_CANDIDATES,
            wanted_ids=wanted_ids,
            completed_before=completed_before,
        )
        print(
            f"ali_api_detail_plan stream=true completed_existing={len(completed_before)} "
            f"limit={args.limit} workers={args.workers}",
            flush=True,
        )

    max_pending = max(1, args.workers) * 8
    with ThreadPoolExecutor(max_workers=max(1, args.workers)) as executor:
        future_map: Dict[Any, Dict[str, Any]] = {}
        for row in todo_iterable:
            item_id = normalize_id(row.get("标的物ID"))
            if not item_id:
                continue
            submitted_ids.add(item_id)
            future = executor.submit(
                fetch_detail,
                row,
                not args.no_description,
                not args.no_attachments,
                not args.no_bid,
                args.raw_text_limit,
                args.retries,
                args.timeout,
            )
            future_map[future] = row
            submitted += 1

            if len(future_map) >= max_pending:
                done, _ = wait(future_map.keys(), return_when=FIRST_COMPLETED)
                for done_future in done:
                    base_row = future_map.pop(done_future)
                    handle_future(done_future, base_row)
                    maybe_report()

            if args.export_every and not args.no_export and completed_this_run > 0 and completed_this_run % args.export_every == 0:
                rows = export_excel_streaming(input_path, jsonl_path, output_path, completed_only=not args.include_unfinished_index)
                print(f"export_checkpoint output={output_path} rows={rows}", flush=True)

        for future in as_completed(future_map):
            base_row = future_map[future]
            handle_future(future, base_row)
            maybe_report()

    if wanted_ids_list:
        missing = [item_id for item_id in wanted_ids_list if item_id not in submitted_ids and item_id not in completed_before]
        if missing:
            print(f"ids_missing_or_completed count={len(missing)} ids={','.join(missing[:10])}", flush=True)

    maybe_report(force=True)

    if args.no_export:
        print(f"ali_api_detail_done jsonl={jsonl_path} export_skipped=true", flush=True)
        return

    rows = export_excel_streaming(input_path, jsonl_path, output_path, completed_only=not args.include_unfinished_index)
    print(f"ali_api_detail_done output={output_path} rows={rows} jsonl={jsonl_path}", flush=True)


if __name__ == "__main__":
    main()
